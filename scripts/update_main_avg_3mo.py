#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_main_avg_3mo.py  —  Refresh the Main-warehouse AVG demand used by Re-Stock V2.

WHAT IT DOES
  Reads a Cin7 "Inventory Movement Details" export (Location = Main Warehouse) and
  rewrites the Main columns of public.branch_avg_monthly_sales:
      avg_sales_main    = ceil( (Sale + SaleMultiple qty_out) / months )
      avg_transfer_main = ceil( max(0, StockTransfer out - in)  / months )
      avg_mth_main      = avg_sales_main + avg_transfer_main   ("what LEAVES Main")
  This total is what Re-Stock V2 uses for pickface coverage (capacity_weeks / runway /
  the red<3wk and orange<4wk alerts / the "ideal cap = 4 weeks of demand" suggestion).
  NOTE: it does NOT change restock_qty, which is always max(0, cap_max - pickface_on_hand).

RULES (agreed with Joao 2026-07-30)
  * Total = what leaves Main = sales + transfer NET (NOT sales-only).
  * ROUND UP (ceil) everything -> whole integers only, never decimals.
  * Products with NO Main movement in the report window are set to 0 (their true
    N-month average is 0 — leaving stale values would mix periods and reintroduce decimals).
  * Case-insensitive product matching: updates the EXISTING row for a SKU even if the
    report's letter-case differs (avoids creating case-duplicate rows).

REPORT
  Cin7 -> Reports -> Product Transaction / "Inventory Movement Details",
  Location = Main Warehouse, any period. Export to xlsx. Header row must be:
      SKU | Category | Reference type | Quantity in | Quantity out | Cost in | Cost out
  Parsed BY HEADER (column order has changed across Cin7 versions — never assume index).

USAGE
  python scripts/update_main_avg_3mo.py                 # dry-run, auto-finds newest report in Downloads
  python scripts/update_main_avg_3mo.py --write         # execute the write
  python scripts/update_main_avg_3mo.py "C:/path/report.xlsx" --write

  Needs SUPABASE_SERVICE_KEY (or SUPABASE_ANON_KEY) in ../.env.
  A CSV backup of prior values is written next to the report before any write.
"""
import sys, os, re, csv, json, math, glob, datetime, urllib.request, urllib.parse

HOST = 'iaqnxamnjftwqdbsnfyl.supabase.co'
TABLE = 'branch_avg_monthly_sales'
DAYS_PER_MONTH = 30.4375  # 365.25 / 12


def load_key():
    env = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env')
    vals = {}
    if os.path.exists(env):
        with open(env, encoding='utf-8') as fh:
            for line in fh:
                m = re.match(r'\s*([A-Z_]+)\s*=\s*(.+)\s*$', line)
                if m:
                    vals[m.group(1)] = m.group(2).strip().strip('"').strip("'")
    key = (os.environ.get('SUPABASE_SERVICE_KEY') or os.environ.get('SUPABASE_ANON_KEY')
           or vals.get('SUPABASE_SERVICE_KEY') or vals.get('SUPABASE_ANON_KEY'))
    if not key:
        sys.exit('No SUPABASE_SERVICE_KEY / SUPABASE_ANON_KEY found in env or ../.env')
    return key


def rest_get(key, path):
    out, off = [], 0
    while True:
        sep = '&' if '?' in path else '?'
        url = f'https://{HOST}/rest/v1/{path}{sep}offset={off}&limit=1000'
        req = urllib.request.Request(url, headers={'apikey': key, 'Authorization': 'Bearer ' + key})
        rows = json.load(urllib.request.urlopen(req))
        out += rows
        if len(rows) < 1000:
            break
        off += 1000
    return out


def rest_post(key, rows):
    req = urllib.request.Request(
        f'https://{HOST}/rest/v1/{TABLE}', data=json.dumps(rows).encode(), method='POST',
        headers={'apikey': key, 'Authorization': 'Bearer ' + key, 'Content-Type': 'application/json',
                 'Prefer': 'resolution=merge-duplicates,return=minimal'})
    with urllib.request.urlopen(req) as r:
        return r.status


def find_report(arg):
    if arg and os.path.exists(arg):
        return arg
    dl = os.path.join(os.environ.get('USERPROFILE', os.path.expanduser('~')), 'Downloads')
    cands = glob.glob(os.path.join(dl, 'Inventory Movement Details*.xlsx'))
    if not cands:
        sys.exit('No "Inventory Movement Details*.xlsx" found in Downloads — pass a path.')
    return max(cands, key=os.path.getmtime)


def parse_report(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Sheet'] if 'Sheet' in wb.sheetnames else wb.worksheets[0]
    meta, header, hdr, data = {}, None, None, []
    for row in ws.iter_rows(values_only=True):
        v = list(row)
        first = str(v[0]) if v[0] is not None else ''
        if header is None:
            m = re.match(r'(From|To):\s*(.+)', first)
            if m:
                meta[m.group(1)] = m.group(2).strip()
            if first == 'SKU':
                header = [str(c).strip() if c is not None else '' for c in v]
                hdr = {h: j for j, h in enumerate(header)}
            continue
        if v[0] is None:
            continue
        data.append(v)
    for need in ('SKU', 'Reference type', 'Quantity in', 'Quantity out'):
        if need not in hdr:
            sys.exit(f'Report header missing "{need}" — got {header}')
    return meta, hdr, data


def months_from_meta(meta):
    try:
        frm = datetime.datetime.strptime(meta['From'], '%d-%b-%Y')
        to = datetime.datetime.strptime(meta['To'], '%d-%b-%Y')
    except (KeyError, ValueError):
        sys.exit('Could not read From/To dates from the report header.')
    raw = (to - frm).days / DAYS_PER_MONTH
    # snap to a whole number of months when very close (a "3-month" pull -> exactly 3.0)
    months = round(raw) if abs(raw - round(raw)) < 0.15 else round(raw, 2)
    return meta['From'], meta['To'], months, raw


def main():
    args = [a for a in sys.argv[1:]]
    do_write = '--write' in args
    path_arg = next((a for a in args if not a.startswith('--')), None)
    key = load_key()
    path = find_report(path_arg)
    meta, hdr, data = parse_report(path)
    frm, to, months, raw = months_from_meta(meta)
    print(f'Report : {os.path.basename(path)}')
    print(f'Period : {frm} -> {to}  ({raw:.2f} months raw -> divisor {months})')

    iS, iR, iIn, iOut = hdr['SKU'], hdr['Reference type'], hdr['Quantity in'], hdr['Quantity out']
    agg = {}
    for v in data:
        sku = str(v[iS]).strip() if v[iS] is not None else ''
        if not sku:
            continue
        ref = str(v[iR]).strip() if v[iR] is not None else ''
        qin = float(v[iIn] or 0)
        qout = float(v[iOut] or 0)
        d = agg.setdefault(sku, {'s': 0.0, 'xo': 0.0, 'xi': 0.0})
        if ref in ('Sale', 'SaleMultiple'):
            d['s'] += qout
        elif ref == 'StockTransfer':
            d['xo'] += qout
            d['xi'] += qin

    # existing rows -> canonical case map (avoid case-dup rows)
    existing = rest_get(key, f'{TABLE}?select=product,avg_mth_main,avg_sales_main,avg_transfer_main')
    canon = {}
    cur = {}
    for r in existing:
        pc = r.get('product') or ''
        canon[pc.upper()] = pc          # keep existing letter-case
        cur[pc.upper()] = r

    movers = set()
    payload = []
    for sku, d in agg.items():
        avg_sales = math.ceil(d['s'] / months)
        avg_xfr = math.ceil(max(0.0, d['xo'] - d['xi']) / months)
        total = avg_sales + avg_xfr
        if total <= 0:
            continue
        movers.add(sku.upper())
        product = canon.get(sku.upper(), sku)   # existing case if present, else report case
        payload.append({'product': product, 'avg_mth_main': total,
                        'avg_sales_main': avg_sales, 'avg_transfer_main': avg_xfr})

    # non-movers currently non-zero -> reset to 0 (their N-month avg is 0)
    zeros = []
    for up, r in cur.items():
        if up in movers:
            continue
        if float(r.get('avg_mth_main') or 0) or float(r.get('avg_sales_main') or 0) or float(r.get('avg_transfer_main') or 0):
            zeros.append({'product': r['product'], 'avg_mth_main': 0, 'avg_sales_main': 0, 'avg_transfer_main': 0})

    new_products = [p['product'] for p in payload if p['product'].upper() not in cur]
    print(f'Movers updated : {len(payload)}  (new products: {len(new_products)})')
    print(f'Non-movers zeroed : {len(zeros)}')

    # backup
    bkp = os.path.join(os.path.dirname(path), f'avg_main_backup_{datetime.date.today().isoformat()}.csv')
    with open(bkp, 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh)
        w.writerow(['product', 'old_total', 'old_sales', 'old_xfr', 'new_total', 'new_sales', 'new_xfr'])
        for p in payload + zeros:
            o = cur.get(p['product'].upper(), {})
            w.writerow([p['product'], o.get('avg_mth_main', ''), o.get('avg_sales_main', ''),
                        o.get('avg_transfer_main', ''), p['avg_mth_main'], p['avg_sales_main'], p['avg_transfer_main']])
    print(f'Backup -> {bkp}')

    if not do_write:
        print('\nDRY RUN — pass --write to execute.')
        return

    now = datetime.datetime.now(datetime.timezone.utc).isoformat()
    both = [dict(p, updated_at=now) for p in payload + zeros]
    B = 500
    for i in range(0, len(both), B):
        st = rest_post(key, both[i:i + B])
        print(f'  batch {i // B + 1}: HTTP {st}  ({min(i + B, len(both))}/{len(both)})')
    # verify no decimals remain
    allrows = rest_get(key, f'{TABLE}?select=avg_mth_main')
    nonint = sum(1 for r in allrows if float(r['avg_mth_main'] or 0) != int(float(r['avg_mth_main'] or 0)))
    print(f'Done. Rows: {len(allrows)}  |  avg_mth_main with decimals: {nonint} (want 0)')


if __name__ == '__main__':
    main()
