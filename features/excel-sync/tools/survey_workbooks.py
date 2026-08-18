"""What shape is each branch workbook, really? Read-only, writes nothing.

Seven branches, three tabs each: twenty-one bindings. The README once assumed
"same three shapes" — the survey exists because that is false. `SOH Main` anchors
at A3 in Brisbane and B3 in Coffs, and a binding that trusts the wrong one writes
the SKU column over the quantities.

So this asks each workbook the four questions a binding has to answer, and the
answers come from the file rather than from anybody's memory:

  where does the data start   the anchor. Found by walking down from the header
                              row, not assumed.

  what is the header row      compared cell for cell at delivery time; a shifted
                              column is refused rather than written through.

  which columns are READ      the load-bearing part. `=VLOOKUP(B4,'SOH Main'!A:E,5)`
                              means column E of that tab is read at offset 5 from
                              A. Move a column and every one of those formulas
                              returns the wrong field, silently and with no error
                              cell to notice. This resolves each reference to an
                              absolute column letter and counts them.

  is a formula in our way     a PATCH of values replaces a formula with a number,
                              permanently. Any formula inside the rectangle is a
                              hard stop, not a warning.

    python tools/survey_workbooks.py                    # table + divergences
    python tools/survey_workbooks.py --emit-bindings out/bindings
    python tools/survey_workbooks.py --folder "C:/path/to/library"

Reads the OneDrive-synced copies because they are already on disk and identical
to what Graph would fetch. Nothing here needs a token.
"""
import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

DEFAULT_FOLDER = os.path.expanduser(
    r'~\RapidLED\WorkDocs - Rapid LED - Data\Inventory Management\Inventory Stock Orders')

# We drive three KINDS of tab and nothing else. Everything else in these
# workbooks — the branch master, the weekly order tabs, Print Layout — stays
# hand-maintained by design.
#
# The names are NOT the same across branches: the branch-stock tab is 'SOH Dear'
# in Brisbane and Coffs, 'SOH CNS' in Cairns, 'SOH SC' on the Sunshine Coast and
# 'SOH Sydney' in Sydney. Hard-coding the Brisbane name reported three branches
# as "tab not present" when the tab was there under another name — so match by
# pattern and let the file say which one it is.
KINDS = [
    # kind             matches                                          dataset
    ('branch-stock',   re.compile(r'^SOH\s+(?!Main)\S', re.I),       'stock-level'),
    ('main-stock',     re.compile(r'^SOH\s+Main\s*$', re.I),           'stock-level'),
    ('supplier-stock', re.compile(r'^SOH\s+Sydney\s*$', re.I),         'stock-level'),
    ('branch-sales',   re.compile(r'^Sales\s+MTD\s*$', re.I),          'monthly-sales'),
]

# Hobart and Melbourne each carry a `SOH Sydney` tab holding SYDNEY's stock -
# 1,487 and 1,484 rows, group header "Sydney", read by 51 formulas each.
# Sydney supplies both branches, the same relationship that makes Sydney's own
# Sales MTD cover SYD+MEL+HOB.
#
# It was first classified as a leftover because 51 formula reads is nothing
# beside SOH Dear's 3,018. That was wrong. A tab can be small and still be
# load-bearing, and 'how many formulas read it' measures how much of the
# workbook leans on it, not whether anybody needs it. The right question was
# what the tab is FOR, and its group header answered that from the start.
SUPPLIER_BRANCHES = ('Hobart', 'Melbourne')

# Branch name inside the workbook -> the Cin7 location it is about. Coffs is the
# one already bound; the rest are asserted here and verified against the tab's
# own group header row below, so a wrong guess shows up as a MISMATCH.
# Workbook name -> the Cin7 location name, which is NOT always the same string.
# Read off the published datasets on 2026-08-14; the Sunshine Coast is the one
# that differs and it cost a refused rehearsal to find, which is the rehearsal
# doing its job. Cin7 also carries per-branch Project warehouses
# ('BNE Project', 'CNS Project', 'SYD Project', 'SC- Project Warehouse') that
# are deliberately NOT included here — the hand-verified Coffs binding counted
# the branch alone, and adding project stock is a business decision, not a
# mapping fix. If a branch tab ever looks short, that is the first thing to ask.
BRANCHES = {
    'Brisbane': 'Brisbane',
    'Cairns': 'Cairns',
    'Coffs Harbour': 'Coffs Harbour',
    'Hobart': 'Hobart',
    'Melbourne': 'Melbourne',
    'Sunshine Coast': 'Sunshine Coast Warehouse',
    'Sydney': 'Sydney',
}

# A branch tab that is NOT about one warehouse.
#
# Sydney's `Sales MTD` covers Sydney, Melbourne and Hobart. Cell F1 of that tab
# says so in the operator's own hand — "Include SYD MEL HOB" — and the numbers
# agree: against the last hand-pasted version Sydney alone matches 133 of 211
# SKUs (63.0%), the three together 328 of 334 (98.2%), the six differences
# totalling 22 units.
#
# It is not a quirk. Sydney is being made a SATELLITE branch: it will hold stock
# and distribute to Melbourne and Hobart, so what it must buy is those three
# branches' demand, not its own. Expect this list to matter more over time, and
# expect Sydney's stock tab to eventually need the same treatment — today it is
# Sydney alone and matches at 99.5%, so it is left alone until the operation
# actually changes.
#
# Found by scanning every cell of the driven tabs for operator notes. If another
# turns up, it belongs here rather than in a comment somewhere.
LOCATION_OVERRIDES = {
    ('Sydney', 'branch-sales'): ['Sydney', 'Melbourne', 'Hobart'],
}


def col_to_num(col):
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def num_to_col(n):
    s = ''
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def branch_of(filename):
    stem = os.path.splitext(os.path.basename(filename))[0]
    for name in BRANCHES:
        if stem.lower().startswith(name.lower()):
            return name
    return None


# ─────────────────────────────────────────────────────────────────────────────
# who reads what
# ─────────────────────────────────────────────────────────────────────────────
def read_map(wb, tab):
    """Absolute columns of `tab` that formulas elsewhere actually read.

    Returns {column_letter: {'hits': n, 'via': set_of_lookup_ranges}} plus the
    key column, which is the leftmost of the lookup range — VLOOKUP matches on
    it, so it is load-bearing in a different way from the value columns.
    """
    esc = re.escape(tab)
    pat = re.compile(
        r"VLOOKUP\(\s*[^,]+,\s*'?" + esc + r"'?!\$?([A-Z]{1,3})\$?(?::\$?([A-Z]{1,3})\$?)?"
        r"\s*,\s*(\d+)", re.IGNORECASE)
    cols = defaultdict(lambda: {'hits': 0, 'via': set()})
    keys = Counter()
    other = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or '!' not in v:
                    continue
                if tab.lower() not in v.lower():
                    continue
                found = pat.findall(v)
                if not found:
                    other += 1
                    continue
                for start, end, idx in found:
                    keys[start.upper()] += 1
                    target = num_to_col(col_to_num(start) + int(idx) - 1)
                    cols[target]['hits'] += 1
                    cols[target]['via'].add(f"{start.upper()}:{(end or start).upper()},{idx}")
    return cols, keys, other


# ─────────────────────────────────────────────────────────────────────────────
# tab shape
# ─────────────────────────────────────────────────────────────────────────────
def describe_tab(wbv, wbf, tab):
    wsv, wsf = wbv[tab], wbf[tab]
    out = {'tab': tab, 'max_row': wsv.max_row, 'max_col': wsv.max_column}

    # Header row = the first row in the top 8 whose cells look like labels and
    # which is followed by data. 'SKU' anchors it in all three tab shapes.
    header_row = None
    key_col = None
    for r in range(1, min(9, wsv.max_row + 1)):
        for c in range(1, min(wsv.max_column, 12) + 1):
            v = wsv.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == 'SKU':
                header_row, key_col = r, c
                break
        if header_row:
            break
    out['header_row'] = header_row
    out['key_col'] = num_to_col(key_col) if key_col else None

    if not header_row:
        out['error'] = "no 'SKU' header found in the first 8 rows"
        return out

    out['headers'] = []
    for c in range(key_col, wsv.max_column + 1):
        v = wsv.cell(header_row, c).value
        if v in (None, ''):
            if c > key_col and all(wsv.cell(header_row, cc).value in (None, '')
                                   for cc in range(c, min(c + 3, wsv.max_column + 1))):
                break
            out['headers'].append((num_to_col(c), None))
            continue
        out['headers'].append((num_to_col(c), str(v).strip()))

    # Anchor: first row under the header carrying a key.
    #
    # An EMPTY tab has no key to find, and the first version of this emitted a
    # literal "None" into the binding — a file that parses and dies at delivery.
    # Melbourne's branch-stock tab is exactly that case: 0 rows, while 6 096
    # formulas read column F of it and every one falls through to ISNA→0. So
    # fall back to the row under the header, which is where the data belongs,
    # and mark the binding so nobody enables it without looking.
    anchor_row = None
    for r in range(header_row + 1, min(header_row + 6, wsv.max_row + 1)):
        if wsv.cell(r, key_col).value not in (None, ''):
            anchor_row = r
            break
    out['anchor_inferred'] = anchor_row is None
    if anchor_row is None:
        anchor_row = header_row + 1
    out['anchor'] = f'{num_to_col(key_col)}{anchor_row}'

    # Extent: last row with a key.
    last = anchor_row or header_row
    for r in range(wsv.max_row, header_row, -1):
        if wsv.cell(r, key_col).value not in (None, ''):
            last = r
            break
    out['data_rows'] = (last - anchor_row + 1) if anchor_row else 0
    out['last_row'] = last

    # The group header above the columns names the Cin7 location this tab is
    # about ('Brisbane', or 'Grand Total' for the main-warehouse tab).
    labels = Counter()
    for r in range(1, header_row):
        for c in range(1, wsv.max_column + 1):
            v = wsv.cell(r, c).value
            if isinstance(v, str) and v.strip():
                labels[v.strip()] += 1
    out['group_labels'] = labels.most_common(6)

    # Formulas inside the rectangle we would write = hard stop.
    formulas = []
    if anchor_row:
        wide = len([h for h in out['headers']])
        for r in range(anchor_row, last + 1):
            for c in range(key_col, key_col + wide):
                v = wsf.cell(r, c).value
                if isinstance(v, str) and v.startswith('='):
                    formulas.append(wsf.cell(r, c).coordinate)
                    if len(formulas) > 40:
                        break
            if len(formulas) > 40:
                break
    out['formulas_in_rect'] = formulas

    # Where a human reads "is this fresh?". Today it is typed by hand.
    stamp = []
    for r in range(1, header_row + 1):
        for c in range(1, wsv.max_column + 1):
            v = wsv.cell(r, c).value
            if isinstance(v, str) and re.search(r'updated', v, re.IGNORECASE):
                stamp.append((wsv.cell(r, c).coordinate, v.strip()))
    out['status_cells'] = stamp

    # Sales MTD carries the report period in its own preamble; it is the cell
    # that contradicted the hand-typed stamp on Brisbane.
    period = []
    for r in range(1, header_row):
        v = wsv.cell(r, 1).value
        if isinstance(v, str) and re.match(r'\s*(report period|from|to|currency)', v, re.IGNORECASE):
            period.append((wsv.cell(r, 1).coordinate, v.strip()))
    out['period_cells'] = period
    # Contiguous only: clearing a range that skips a used cell would blank it.
    out['clear_range'] = None
    if period:
        rows_ = [int(c[1:]) for c, _ in period]
        if rows_ == list(range(min(rows_), max(rows_) + 1)):
            out['clear_range'] = f'A{min(rows_)}:A{max(rows_)}'

    # Where the stamp and the four-line block go.
    #
    # Prefer the cell somebody already types "Updated: 10-Aug" into — people
    # look there, and replacing it in place is the whole point. When a tab has
    # no such cell, DO NOT fall back to a fixed guess: 'H1' happens to hold a
    # leftover header on several of these tabs, and a default that silently
    # lands on real content is worse than no default. Find an empty column
    # clear of the data instead, and say that is what happened.
    data_right = key_col + len(out['headers']) - 1

    def free_block(scol, brow=3, w=2, h=4):
        return not [1 for r in range(brow, brow + h) for c in range(scol, scol + w)
                    if wsv.cell(r, c).value not in (None, '')]

    out['status_block'] = None
    out['status_cell'] = None
    out['status_cell_source'] = None

    if stamp:
        out['status_cell'] = stamp[0][0]
        out['status_cell_source'] = 'existing hand-typed stamp'
        scol = col_to_num(re.match(r'([A-Z]+)', stamp[0][0]).group(1))
        if scol <= data_right:
            out['status_block_blocked_by'] = f'column {num_to_col(scol)} is inside the data'
        elif free_block(scol):
            out['status_block'] = f'{num_to_col(scol)}3'
        else:
            out['status_block_blocked_by'] = [
                wsv.cell(r, c).coordinate for r in range(3, 7) for c in (scol, scol + 1)
                if wsv.cell(r, c).value not in (None, '')]
    else:
        for scol in range(data_right + 2, data_right + 14):
            if wsv.cell(1, scol).value in (None, '') and free_block(scol):
                out['status_cell'] = f'{num_to_col(scol)}1'
                out['status_cell_source'] = 'chosen — tab had no stamp; first free column right of the data'
                out['status_block'] = f'{num_to_col(scol)}3'
                break
        if not out['status_cell']:
            out['status_cell_source'] = 'NONE FOUND — no empty column within 12 of the data'
    return out


def survey_file(path):
    wbv = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path, data_only=False)
    rec = {'file': os.path.basename(path), 'branch': branch_of(path),
           'sheets': wbv.sheetnames, 'tabs': []}

    for kind, pattern, dataset in KINDS:
        # Sydney's own workbook uses `SOH Sydney` as its BRANCH tab, so it must
        # not also be bound as a supplier tab.
        if kind == 'supplier-stock' and rec['branch'] not in SUPPLIER_BRANCHES:
            continue
        if kind == 'branch-stock' and rec['branch'] in SUPPLIER_BRANCHES:
            pattern = re.compile(r'^SOH\s+(?!Main|Sydney)\S', re.I)
        matches = [s for s in wbv.sheetnames if pattern.match(s.strip())]
        if not matches:
            rec['tabs'].append({'kind': kind, 'error': 'NO TAB MATCHES', 'dataset': dataset})
            continue
        # More than one candidate happens: Hobart and Melbourne carry both a
        # 'SOH Dear' and a 'SOH Sydney'. Describe them all and let the formula
        # count say which one is actually driving the order sheet — a tab that
        # nothing reads is a leftover, whatever it is called.
        for name in matches:
            d = describe_tab(wbv, wbf, name)
            d['kind'] = kind
            d['dataset'] = dataset
            d['actual_name'] = name
            cols, keys, other = read_map(wbf, name)
            d['read_cols'] = {k: {'hits': v['hits'], 'via': sorted(v['via'])}
                              for k, v in sorted(cols.items())}
            d['lookup_key_cols'] = dict(keys)
            d['non_vlookup_refs'] = other
            d['total_reads'] = sum(v['hits'] for v in cols.values()) + other
            d['ambiguous'] = len(matches) > 1
            rec['tabs'].append(d)
    return rec


def driven(rec, kind):
    """The tab of this kind the workbook actually depends on: most-read wins."""
    cands = [t for t in rec['tabs'] if t.get('kind') == kind and not t.get('error')]
    if not cands:
        return next((t for t in rec['tabs'] if t.get('kind') == kind), None)
    return max(cands, key=lambda t: t.get('total_reads', 0))


# ─────────────────────────────────────────────────────────────────────────────
# output
# ─────────────────────────────────────────────────────────────────────────────
def print_report(recs):
    print('=' * 112)
    print('PER-BRANCH TAB SHAPE  —  the tab this workbook actually depends on, per kind'.center(112))
    print('=' * 112)
    print(f"{'branch':<16}{'kind':<14}{'tab name':<13}{'anchor':>7}{'hdr':>5}{'rows':>7}{'key':>5}"
          f"  {'READ by':<14}{'formulas in rect':>18}")
    print('-' * 112)
    for r in recs:
        for kind, _, _ in KINDS:
            d = driven(r, kind)
            if not d or d.get('error'):
                print(f"{r['branch']:<16}{kind:<14}{(d or {}).get('error', 'MISSING')}")
                continue
            reads = ' '.join(f"{c}({v['hits']})" for c, v in d['read_cols'].items()) or '(none)'
            nf = len(d['formulas_in_rect'])
            flag = 'NONE — safe' if nf == 0 else f'*** {nf}+ BLOCKS ***'
            warn = '  <- EMPTY!' if d['data_rows'] == 0 else ''
            print(f"{r['branch']:<16}{kind:<14}{d['actual_name']:<13}{d['anchor'] or '?':>7}"
                  f"{d['header_row'] or '?':>5}{d['data_rows']:>7}{d['key_col'] or '?':>5}"
                  f"  {reads:<14}{flag:>18}{warn}")
    print()
    extra = [(r['branch'], t) for r in recs for t in r['tabs']
             if t.get('ambiguous') and t is not driven(r, t['kind'])]
    if extra:
        print('  Also present, but read by fewer formulas — leftovers, do not bind:')
        for br, t in extra:
            print(f"    {br:<16}{t['actual_name']:<13}{t.get('data_rows', 0):>6} rows, "
                  f"{t.get('total_reads', 0)} formula reads"
                  + (f"   [{t['error']}]" if t.get('error') else ''))
        print()


def print_divergence(recs):
    print('=' * 112)
    print('DIVERGENCE — where the branches disagree (this is what breaks a copy-pasted binding)'.center(112))
    print('=' * 112)
    for kind, _, _ in KINDS:
        groups = defaultdict(list)
        for r in recs:
            d = driven(r, kind)
            if not d or d.get('error'):
                groups[f"ERROR: {(d or {}).get('error', 'MISSING')}"].append(r['branch'])
                continue
            sig = (d['actual_name'], d['anchor'], d['header_row'],
                   tuple(h for _, h in d['headers']), tuple(sorted(d['read_cols'])))
            groups[sig].append(r['branch'])
        print(f'\n  {kind}  —  {len(groups)} distinct shape(s)')
        for sig, branches in sorted(groups.items(), key=lambda kv: -len(kv[1])):
            if isinstance(sig, str):
                print(f'    [{len(branches)}] {", ".join(branches)}  ->  {sig}')
                continue
            name, anchor, hrow, headers, reads = sig
            print(f'    [{len(branches)}] {", ".join(branches)}')
            print(f'         tab {name!r} · anchor {anchor} · header row {hrow} · reads {" ".join(reads) or "(none)"}')
            print(f'         headers: {" | ".join(h or "·" for h in headers)}')
    print()


def print_detail(recs):
    print('=' * 112)
    print('LOAD-BEARING COLUMNS — resolved from every VLOOKUP that points at these tabs'.center(112))
    print('=' * 112)
    for r in recs:
        print(f"\n  {r['branch']}  ({r['file']})")
        for kind, _, _ in KINDS:
            d = driven(r, kind)
            if not d or d.get('error'):
                print(f'    {kind:<14} {(d or {}).get("error", "MISSING")}')
                continue
            print(f'    {kind:<14} {d["actual_name"]!r}  anchor {d["anchor"]}  ·  '
                  f'{d["data_rows"]} rows  ·  key col {d["key_col"]}')
            if d['read_cols']:
                for col, v in d['read_cols'].items():
                    label = next((h for cl, h in d['headers'] if cl == col), None)
                    print(f'         col {col:<3} "{label}"  <- {v["hits"]} formulas  via {", ".join(v["via"])}')
            else:
                print('         (no formula reads this tab — nothing downstream depends on it)')
            if d['status_cells']:
                for coord, txt in d['status_cells']:
                    print(f'         stamp {coord}: {txt!r}   (hand-typed today)')
            if d['period_cells']:
                for coord, txt in d['period_cells']:
                    print(f'         period {coord}: {txt!r}')
            if d['formulas_in_rect']:
                print(f'         !! FORMULAS INSIDE WRITE RECTANGLE: {", ".join(d["formulas_in_rect"][:12])}')
    print()


TOML = '''slug = "{slug}"
title = "{branch} — {tab}"
dataset = "{dataset}"
what_it_does = "{what}"
feeds = ["Order Template", "weekly order tabs"]

# ── Scheduling and the on/off switch ────────────────────────────────────────
# These live HERE, above every [section], and moving them is a real bug rather
# than a style choice. In TOML every key after a [section] header belongs to
# that section, so with these at the bottom of the file they parsed as
# status.enabled and status.sla_minutes — and binding['enabled'] came back None.
#
# What that cost: setting enabled = true would have done NOTHING. The binding
# would have stayed disabled, delivery would have skipped it, the registry would
# have shown it off, and the heartbeat would not have watched it — while whoever
# flipped the flag believed the branch was live. Going live is exactly when that
# lie would have been told.
cron_utc = "0 19 * * 0-4"
sla_minutes = 1560
enabled = false

# Generated by tools/survey_workbooks.py from the live workbook. Every value
# below was read out of the file, not assumed. Re-run the survey after any
# change to the tab and diff this file; that is cheaper than discovering a
# moved column at 07:00.

[workbook]
# The seven workbooks sit side by side in one folder. Historically each was
# renamed every month ("Brisbane Aug 26" -> "Brisbane Sep 26") and the old one
# moved into `2026/<Mon YY>/`, which no binding can follow. The agreed fix is
# the other way round: this file keeps ONE name for good, and the monthly
# archive copy is the one that gets renamed.
#
# So if the name below stops resolving, that is a person renaming the live file,
# and delivery failing loudly is the correct outcome — not something to paper
# over with a wildcard.
file = "{file}"
sheet = "{tab}"
library = "Rapid LED - Data"
folder = "Inventory Management/Inventory Stock Orders"

[layout]
kind = "flat"
key = "sku"
data_anchor = "{anchor}"          # header row {header_row}; {rows} data rows at survey time{anchor_warn}
locations = [{locations}]

{columns}
[status]
# cell = the one-line stamp, written where the hand-typed one already sits so
# people keep looking in the same place. block = the four lines under it:
# Updated / Covers / Rows / Source. `Covers` is the answer to the question the
# stamp alone never answered — Brisbane showed June figures under an
# "Updated 10-Aug" stamp because the two came from different hands.
cell = "{status_cell}"{status_note}{status_block}{status_clear}
'''


def emit_bindings(recs, outdir):
    os.makedirs(outdir, exist_ok=True)
    written = []
    for r in recs:
        for kind, _, _ in KINDS:
            d = driven(r, kind)
            if not d or d.get('error'):
                continue
            slug = r['branch'].lower().replace(' ', '-') + '-' + kind
            cols = []
            for col, label in d['headers']:
                if not label:
                    continue
                field = {'SKU': 'sku', 'Quantity on hand': 'on_hand', 'Allocated': 'allocated',
                         'On order': 'on_order', 'Available': 'available', 'Unit': 'unit',
                         'Quantity': 'quantity', 'Discount': 'discount', 'Total': 'total',
                         'Unit cost': 'unit_cost', 'In transit': 'in_transit',
                         'Stock on hand': 'stock_value'}.get(label)
                if not field:
                    cols.append(f'# column {col} "{label}" — no dataset field; left out on purpose')
                    continue
                note = ''
                if col in d['read_cols']:
                    v = d['read_cols'][col]
                    note = (f'\n# Column {col}. READ BY {v["hits"]} FORMULAS via {", ".join(v["via"])}.\n'
                            f'# Its position inside that range is load-bearing; do not reorder.')
                extra = ''
                if field == 'discount':
                    # Not in the dataset — the aggregation is ambiguous and no
                    # formula reads it. The column stays so the layout still
                    # matches the Cin7 export it replaces.
                    extra = '\nblank = true'
                    note = note or '\n# Dropped from the dataset; written blank to hold the layout.'
                elif field in ('total', 'stock_value'):
                    extra = '\nround = 2'
                cols.append(f'[[columns]]{note}\nfield = "{field}"\nheader = "{label}"{extra}')
            override = LOCATION_OVERRIDES.get((r['branch'], kind))
            # A supplier tab is about the SUPPLIER, not the branch whose workbook
            # it sits in. Defaulting to the branch put Hobart's stock into a tab
            # headed "Sydney" — the tab's own group header row said so and the
            # generator ignored it.
            locs = (', '.join(f'"{x}"' for x in override) if override
                    else '"Main Warehouse", "Gateway"' if kind == 'main-stock'
                    else '"Sydney"' if kind == 'supplier-stock'
                    else f'"{BRANCHES[r["branch"]]}"')
            what = ('Main Warehouse + Gateway stock on hand, summed.' if kind == 'main-stock'
                    else f'{r["branch"]} stock on hand.' if kind == 'branch-stock'
                    else f'{r["branch"]} sales for the current month, SKU by SKU.')
            body = TOML.format(
                slug=slug, branch=r['branch'], tab=d['actual_name'],
                dataset=d['dataset'], what=what, file=r['file'],
                anchor=d['anchor'], header_row=d['header_row'], rows=d['data_rows'],
                anchor_warn=('' if not d.get('anchor_inferred') else
                             '\n# ⚠ INFERRED, NOT MEASURED — this tab is EMPTY today, so there was no\n'
                             '# first data row to find; the anchor is just "the row under the header".\n'
                             '# Confirm it against the tab before enabling. An empty tab that formulas\n'
                             '# still read is a live problem in its own right, separate from this sync.'),
                locations=locs, columns='\n\n'.join(cols) + '\n',
                status_cell=d.get('status_cell') or 'H1',
                status_note='' if d.get('status_cell_source', '').startswith('existing')
                else f'\n# {d.get("status_cell_source")}',
                status_block=(f'\nblock = "{d["status_block"]}"' if d.get('status_block') else
                              '\n# block = "??"   no empty 2x4 found near the stamp'
                              f' ({d.get("status_block_blocked_by")})'),
                status_clear=(f'\nclear = "{d["clear_range"]}"'
                              '     # the export preamble, whose period goes stale'
                              if d.get('clear_range') else ''))
            path = os.path.join(outdir, slug + '.toml')
            with open(path, 'w', encoding='utf-8') as fh:
                fh.write(body)
            written.append(path)
    print(f'  wrote {len(written)} draft bindings to {outdir}')
    print('  every one is enabled = false. Review before flipping any of them.')


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--folder', default=DEFAULT_FOLDER)
    ap.add_argument('--emit-bindings', metavar='DIR')
    ap.add_argument('--json', metavar='FILE')
    args = ap.parse_args()

    if not os.path.isdir(args.folder):
        raise SystemExit(f'folder not found: {args.folder}')
    files = sorted(f for f in os.listdir(args.folder)
                   if f.lower().endswith('.xlsx') and not f.startswith('~$') and branch_of(f))
    if not files:
        raise SystemExit(f'no branch workbooks in {args.folder}')

    print(f'folder  {args.folder}')
    print(f'files   {len(files)}: {", ".join(files)}\n')
    print('This survey READS ONLY. It opens no connection and writes to no workbook.\n')

    recs = []
    for f in files:
        recs.append(survey_file(os.path.join(args.folder, f)))

    print_report(recs)
    print_divergence(recs)
    print_detail(recs)

    if args.json:
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(recs, fh, indent=1, default=str)
        print(f'  json -> {args.json}')
    if args.emit_bindings:
        emit_bindings(recs, args.emit_bindings)
    return 0


if __name__ == '__main__':
    sys.exit(main())
