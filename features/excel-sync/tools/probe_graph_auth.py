"""Which Microsoft Graph door is actually open, all the way to a written cell?

Six tests, and only the sixth settles anything:

    1  authenticate            2  reach the site       3  find the workbook
    4  createSession           5  GET a range          6  PATCH a range

Tests 1-3 pass under app-only. So does finding the file, reading its size, and
listing its worksheets. Everything looks configured — and then the first real
write returns 403, because Microsoft lists Application as **Not supported** for
both `createSession` and range update. The permission tables were verified on
2026-08-13; delegated `Files.ReadWrite` is the only entry.

That is the trap this probe exists to spring early. Asking an administrator for
app-only `Sites.Selected` and expecting it to unblock workbook writes costs a
full consent cycle to discover, at the point where twenty-one bindings are
already configured and the failure looks like a bug in our code.

  A) app-only    client credentials + `Sites.Selected`. Correct for driveItem
                 calls. Cannot edit cells. Run with --write-test to watch it
                 fail at step 6 — that output is the evidence for the IT ticket.
  B) delegated   device code once, then a refresh token that renews itself on
                 every run. The only door that reaches step 6.

    python tools/probe_graph_auth.py                     # tests 1-3, reads only
    python tools/probe_graph_auth.py --write-test        # tests 1-6
    python tools/probe_graph_auth.py --delegated --write-test

--write-test creates and deletes one throwaway workbook in the library. It never
touches a branch file: a PATCH is the one call that cannot be undone, so it is
proven somewhere disposable or not at all.

Device code needs "Allow public client flows = Yes" on the app registration —
that switch belongs to whoever owns the app, not to a tenant administrator.
"""
import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request

# This runs from a Windows console, where the default cp1252 cannot encode the
# bullets below — it would die on a UnicodeEncodeError instead of printing the
# one answer it exists to give.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

# Not secrets. The status doc carries them in the open for exactly this reason.
TENANT = '59ec4380-0cab-455d-a6e2-f10314801005'
CLIENT = '8c4aa84e-db46-4d6c-b629-922e7ca22243'

# The Coffs workbook, by path. Located 2026-08-11 in the synced library, which is
# more reliable than the sharing URL: these files are renamed every month, and a
# path survives that where a captured document ID does not.
HOST = 'rapidled.sharepoint.com'
SITE_PATH = ''                                   # '' = the root site
LIBRARY = 'Rapid LED - Data'
FOLDER = 'Inventory Management/Inventory Stock Orders'
FILENAME = 'Coffs Harbour Aug 26.xlsx'

GRAPH = 'https://graph.microsoft.com/v1.0'
DELEGATED_SCOPES = 'https://graph.microsoft.com/Files.ReadWrite.All offline_access'


def _env(name):
    """Real env wins; fall back to the repo .env, same order as engine/supabase.py."""
    if os.environ.get(name):
        return os.environ[name]
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
    path = os.path.join(root, '.env')
    if os.path.exists(path):
        with open(path, encoding='utf-8') as fh:
            for line in fh:
                m = re.match(r'\s*([A-Z0-9_]+)\s*=\s*(.*)\s*$', line)
                if m and m.group(1) == name:
                    return m.group(2).strip().strip('"').strip("'")
    return None


def _post(url, form):
    body = urllib.parse.urlencode(form).encode()
    req = urllib.request.Request(url, data=body, method='POST')
    req.add_header('Content-Type', 'application/x-www-form-urlencoded')
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        raw = e.read()
        try:
            return e.code, json.loads(raw)
        except ValueError:
            return e.code, {'raw': raw.decode('utf-8', 'replace')[:400]}


def _get(url, token):
    return _call('GET', url, token)


def _call(method, url, token, body=None, raw_body=None, content_type=None, session=None):
    req = urllib.request.Request(url, method=method)
    req.add_header('Authorization', f'Bearer {token}')
    if session:
        req.add_header('workbook-session-id', session)
    data = None
    if raw_body is not None:
        data = raw_body
        req.add_header('Content-Type', content_type or 'application/octet-stream')
    elif body is not None:
        data = json.dumps(body).encode()
        req.add_header('Content-Type', 'application/json')
    if data is not None:
        req.data = data
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            payload = r.read()
            if not payload:
                return r.status, {}
            try:
                return r.status, json.loads(payload)
            except ValueError:
                return r.status, {'raw': payload.decode('utf-8', 'replace')[:400]}
    except urllib.error.HTTPError as e:
        raw = e.read()
        try:
            return e.code, json.loads(raw)
        except ValueError:
            return e.code, {'raw': raw.decode('utf-8', 'replace')[:400]}


def _err(data):
    e = data.get('error') or {}
    if isinstance(e, dict):
        return f"{e.get('code', '?')} — {str(e.get('message', data))[:200]}"
    return str(data)[:200]


def find_workbook(token):
    """Resolve host -> site -> library -> file. Returns (item, drive_id) or (None, why).

    Kept separate from the auth checks on purpose: a token that works and a file
    that cannot be found are two different answers, and collapsing them would
    report "no door is open" when the only problem is a wrong path.
    """
    site_ref = f'{HOST}:{SITE_PATH}' if SITE_PATH else HOST
    st, site = _get(f'{GRAPH}/sites/{site_ref}?$select=id,displayName', token)
    if st != 200:
        return None, f"site {site_ref}: {(site.get('error') or {}).get('message', st)}"

    st, drives = _get(f"{GRAPH}/sites/{site['id']}/drives?$select=id,name", token)
    if st != 200:
        return None, f"drives: {(drives.get('error') or {}).get('message', st)}"
    match = [d for d in drives.get('value', []) if d.get('name') == LIBRARY]
    if not match:
        names = ', '.join(sorted(d.get('name', '?') for d in drives.get('value', [])))
        return None, f'library {LIBRARY!r} not found. Present: {names}'
    drive_id = match[0]['id']

    rel = urllib.parse.quote(f'{FOLDER}/{FILENAME}')
    st, item = _get(f'{GRAPH}/drives/{drive_id}/root:/{rel}'
                    '?$select=id,name,size,lastModifiedDateTime', token)
    if st != 200:
        return None, f"file: {(item.get('error') or {}).get('message', st)}"
    return (item, drive_id), None


PROBE_FILE = '_excel_sync_probe_delete_me.xlsx'


def check_write_path(token, drive_id, label):
    """Tests 4-6: createSession, GET range, PATCH range — the ones that decide.

    Runs against a throwaway workbook this function creates and deletes, never
    against a branch file. That matters: a PATCH is the one call that cannot be
    undone, and the whole point of the probe is to find out whether it works
    BEFORE any real workbook is configured.

    Tests 1-3 (token, site, file) pass under app-only as well, which is exactly
    how a dead end gets mistaken for a working setup. Only the PATCH separates
    them.
    """
    print(f'\n  --- write path ({label}) ---')
    try:
        import io
        import openpyxl as _x
    except ImportError:
        print('    SKIP — openpyxl not installed (pip install -r requirements.txt)')
        return None

    wb = _x.Workbook()
    ws = wb.active
    ws.title = 'Probe'
    ws['A1'], ws['B1'] = 'probe', 'before'
    buf = io.BytesIO()
    wb.save(buf)

    rel = urllib.parse.quote(f'{FOLDER}/{PROBE_FILE}')
    base = f'{GRAPH}/drives/{drive_id}/root:/{rel}:'

    st, item = _call('PUT', base + '/content', token, raw_body=buf.getvalue(),
                     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if st not in (200, 201):
        print(f'    4-6  upload FAILED {st}: {_err(item)}')
        print('         cannot test writing without somewhere safe to write.')
        return False
    item_id = item['id']
    print(f'    ok   uploaded {PROBE_FILE}')
    wbk = f'{GRAPH}/drives/{drive_id}/items/{item_id}/workbook'
    result = False
    session = None
    try:
        st, s = _call('POST', wbk + '/createSession', token, body={'persistChanges': True})
        if st in (200, 201) and s.get('id'):
            session = s['id']
            print('    4    createSession           OK')
        else:
            print(f'    4    createSession           FAILED {st}: {_err(s)}')

        st, rng = _call('GET', wbk + "/worksheets('Probe')/range(address='A1:B1')",
                        token, session=session)
        if st == 200:
            print(f"    5    GET range               OK   values={rng.get('values')}")
        else:
            print(f'    5    GET range               FAILED {st}: {_err(rng)}')

        st, out = _call('PATCH', wbk + "/worksheets('Probe')/range(address='A1:B1')",
                        token, body={'values': [['probe', 'AFTER']]}, session=session)
        if st == 200:
            st2, back = _call('GET', wbk + "/worksheets('Probe')/range(address='A1:B1')",
                              token, session=session)
            got = (back.get('values') or [[None, None]])[0][1] if st2 == 200 else None
            if got == 'AFTER':
                print('    6    PATCH range             OK   <<< the decisive result')
                result = True
            else:
                print(f'    6    PATCH range             returned 200 but read back {got!r}')
        else:
            print(f'    6    PATCH range             FAILED {st}: {_err(out)}')
            if st in (401, 403):
                print('         If this is app-only: expected. Microsoft lists Application as')
                print('         "Not supported" for range update and createSession. No amount of')
                print('         consent changes it — the workbook API needs a delegated token.')
    finally:
        if session:
            _call('POST', wbk + '/closeSession', token, body={}, session=session)
        st, _ = _call('DELETE', f'{GRAPH}/drives/{drive_id}/items/{item_id}', token)
        print(f"    ok   probe file deleted" if st in (200, 204)
              else f'    !!   COULD NOT DELETE {PROBE_FILE} — remove it by hand ({st})')
    return result


def check_app(write_test=False):
    """Did the admin consent land? One call beats staring at a greyed-out button."""
    print('\n=== A) app-only (client credentials) ===')
    secret = _env('GRAPH_CLIENT_SECRET')
    if not secret:
        print('  SKIP — GRAPH_CLIENT_SECRET not in env or .env.')
        print('  Without it this check cannot run; it does not mean consent is missing.')
        return None

    status, data = _post(
        f'https://login.microsoftonline.com/{TENANT}/oauth2/v2.0/token',
        {'client_id': CLIENT, 'client_secret': secret,
         'scope': 'https://graph.microsoft.com/.default',
         'grant_type': 'client_credentials'},
    )
    if status == 200 and data.get('access_token'):
        # A token proves the secret works. It does NOT prove Sites.Selected was
        # consented, nor that any site was granted — those fail later, at the
        # first real call. So make one.
        print('  token: OK')
        found, why = find_workbook(data['access_token'])
        if found:
            item, drive = found
            print(f"  READ OK — {item.get('name')} · {item.get('size')} bytes · "
                  f"modified {item.get('lastModifiedDateTime')}")
            print('  => consent AND a site grant are in place: this token can reach the file.')
            print('     That is NOT the same as being able to edit cells — see the write path.')
            wrote = check_write_path(data['access_token'], drive, 'app-only') if write_test else None
            if wrote:
                print('  => app-only can write ranges. That contradicts the documented')
                print('     permission table — re-read it before relying on this.')
            elif wrote is False:
                print('  => app-only reads but cannot write. Expected. Delivery must use delegated.')
            return True
        print(f'  read FAILED: {why}')
        print('  => the secret works, but the app cannot reach the file. Either the')
        print('     Sites.Selected consent or the per-site grant is still missing.')
        return False

    err = (data.get('error_description') or data.get('error') or data)
    print(f'  token FAILED {status}: {str(err)[:300]}')
    return False


def check_delegated(write_test=False):
    """Sign in as a human. No administrator involved, if the tenant permits it."""
    print('\n=== B) delegated (device code, as the signed-in user) ===')
    status, dc = _post(
        f'https://login.microsoftonline.com/{TENANT}/oauth2/v2.0/devicecode',
        {'client_id': CLIENT, 'scope': DELEGATED_SCOPES},
    )
    if status != 200 or 'device_code' not in dc:
        err = dc.get('error_description') or dc.get('error') or dc
        print(f'  device code FAILED {status}: {str(err)[:400]}')
        if 'AADSTS7000218' in str(err) or 'client_assertion' in str(err):
            print('  => turn ON "Allow public client flows" in the app registration')
            print('     (Authentication → Advanced settings). Joao owns the app; no admin needed.')
        return False

    print(f"\n  {dc['message']}\n")
    print('  waiting for sign-in…')
    deadline = time.time() + int(dc.get('expires_in', 900))
    interval = int(dc.get('interval', 5))
    token = None
    while time.time() < deadline:
        time.sleep(interval)
        st, tk = _post(
            f'https://login.microsoftonline.com/{TENANT}/oauth2/v2.0/token',
            {'client_id': CLIENT, 'grant_type': 'urn:ietf:params:oauth:grant-type:device_code',
             'device_code': dc['device_code']},
        )
        if st == 200:
            token = tk
            break
        code = tk.get('error')
        if code == 'authorization_pending':
            continue
        if code == 'slow_down':
            interval += 5
            continue
        desc = tk.get('error_description', '')
        print(f'  sign-in FAILED: {code} — {str(desc)[:400]}')
        if 'AADSTS65001' in desc or 'consent' in desc.lower():
            print('  => the tenant blocks user consent for this app. Delegated is out')
            print('     too, and admin consent is the only remaining door.')
        return False

    if not token:
        print('  timed out waiting for sign-in.')
        return False

    # THE answer. Everything below is about the file, not about the door.
    print('  token: OK  <<< delegated auth works — this is the decisive result')
    if token.get('refresh_token'):
        print('  refresh_token: present — a nightly headless run can renew itself.')
    else:
        print('  refresh_token: MISSING — add offline_access to the scopes.')

    found, why = find_workbook(token['access_token'])
    if not found:
        print(f'  file NOT reachable: {why}')
        print('  => auth is fine; the path is wrong or that library is not shared')
        print('     with this account. Fixable in the binding — not a blocker.')
        return True

    item, drive = found
    print(f"  READ OK — {item.get('name')} · {item.get('size')} bytes · "
          f"modified {item.get('lastModifiedDateTime')}")

    # Reading the tab names proves we can address the workbook itself, not just
    # the file blob — which is what Phase 5 actually needs.
    st, ws = _get(f"{GRAPH}/drives/{drive}/items/{item['id']}/workbook/worksheets",
                  token['access_token'])
    if st == 200:
        names = [w['name'] for w in ws.get('value', [])]
        print(f'  worksheets ({len(names)}): {", ".join(names[:12])}'
              + (' …' if len(names) > 12 else ''))
        # The branch-stock tab is named per branch — 'SOH Dear' here, but
        # 'SOH CNS' in Cairns, 'SOH SC' on the Sunshine Coast, 'SOH Sydney' in
        # Sydney. tools/survey_workbooks.py resolves that per file.
        for wanted in ('SOH Main', 'Sales MTD'):
            print(f"    {wanted:<10} {'found' if wanted in names else 'NOT FOUND'}")
        branch_tab = [n for n in names if n.upper().startswith('SOH ')
                      and n.strip().upper() != 'SOH MAIN']
        print(f"    branch SOH {branch_tab or 'NOT FOUND'}")
    else:
        err = (ws.get('error') or {})
        print(f"  worksheets FAILED {st}: {err.get('code')} — {str(err.get('message'))[:250]}")
        print('  => file is reachable but the workbook API is not. Worth a look.')
        return True

    if write_test:
        wrote = check_write_path(token['access_token'], drive, 'delegated')
        if wrote:
            print('\n  => delegated works END TO END, PATCH included. Delivery can proceed.')
        else:
            print('\n  => reads fine, cannot write. Check the scopes include Files.ReadWrite.All')
            print('     and that this account has edit rights on the library.')
    else:
        print('\n  => delegated reads end to end. The PATCH is UNPROVEN — rerun with')
        print('     --write-test before configuring any binding.')
    return True


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--app', action='store_true', help='app-only check only')
    ap.add_argument('--delegated', action='store_true', help='delegated check only')
    ap.add_argument('--write-test', action='store_true',
                    help='also run tests 4-6 (createSession, GET range, PATCH range) '
                         'against a throwaway workbook it creates and deletes')
    args = ap.parse_args()

    both = not (args.app or args.delegated)
    print(f'tenant  {TENANT}\nclient  {CLIENT}')
    if args.write_test:
        print(f'Write test ON: creates and deletes {PROBE_FILE} in the library.')
        print('It never touches a branch workbook.')
    else:
        print('This probe READS ONLY. Tests 1-3 (token, site, file) prove nothing about')
        print('writing — they pass on app-only too. Add --write-test for tests 4-6.')

    app = check_app(args.write_test) if (both or args.app) else None
    dele = check_delegated(args.write_test) if (both or args.delegated) else None

    print('\n=== verdict ===')
    if dele:
        print('  delegated works — this is the door delivery must use. Workbook range')
        print('  writes have no app-only path (Application: Not supported), so the')
        print('  refresh token is the destination here, not a bridge.')
        print('  Sign in as a dedicated automation account, not a person: the version')
        print('  history of seven workbooks will carry that name every single day.')
        if app:
            print('  app-only also works for driveItem calls — keep it for those if useful.')
    elif app:
        print('  app-only reaches the file, delegated did not sign in. That is NOT enough:')
        print('  app-only cannot PATCH a range. Get delegated working before building.')
    elif dele is None and app is None:
        print('  nothing was checked. --app needs GRAPH_CLIENT_SECRET; --delegated needs')
        print('  a browser sign-in. Neither ran, so no door has been ruled out.')
    elif dele is None:
        print('  delegated was not checked, and it is the one that matters. Rerun with')
        print('  --delegated --write-test.')
    else:
        print('  both doors shut. Remaining options: an admin consent for delegated, or')
        print('  fall back to writing the OneDrive-synced local copy from a Windows box.')
    if not args.write_test:
        print('\n  The PATCH is still unproven. Rerun with --write-test.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
