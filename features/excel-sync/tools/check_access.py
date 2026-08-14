"""Can the job actually reach, open and write every workbook it is bound to?

Read-only. It opens each file for writing and closes it without writing a byte,
which proves the permission without touching the content — no timestamp moves,
no bytes change. Everything else it checks is a read.

The point is to answer, before anything is enabled, the question that otherwise
gets answered at 07:00 on a Monday: *for all 21 bindings, does the file resolve,
is it ours to write, is the tab still the shape the binding expects, and is
anything in the way?*

    python tools/check_access.py                    # the live SharePoint library
    python tools/check_access.py --root "C:\\...\\Tests files"
    python tools/check_access.py --root "…" --slug brisbane-main-stock

A cloud-only placeholder is called out separately from a missing file: OneDrive
leaves the name visible with no content behind it, and reading it silently
triggers a download that may not finish. The two look identical in Explorer and
fail very differently at 07:00.
"""
import argparse
import ctypes
import os
import sys

import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from engine import pivot                                    # noqa: E402
from engine.delivery import CONFIG, graph                    # noqa: E402

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

FILE_ATTRIBUTE_OFFLINE = 0x1000
FILE_ATTRIBUTE_READONLY = 0x0001
FILE_ATTRIBUTE_RECALL_ON_OPEN = 0x40000
FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS = 0x400000

OK, WARN, BAD = 'ok', 'warn', 'BAD'


def attrs(path):
    try:
        a = ctypes.windll.kernel32.GetFileAttributesW(str(path))
        return None if a == 0xFFFFFFFF else a
    except (AttributeError, OSError):
        return None            # not Windows; the other checks still apply


def resolve(binding, root):
    """Where this binding's workbook lives, trying the folder layout then flat.

    The live library keeps the workbooks under
    `<root>/Inventory Management/Inventory Stock Orders/`. A folder of test
    copies keeps them side by side. Both are legitimate, so try the structured
    path first and fall back — which means a whole run can be pointed at test
    copies with one --root and no per-binding overrides.
    """
    name = binding['workbook']['file']
    folder = (binding['workbook'].get('folder') or '').replace('/', os.sep)
    candidates = [os.path.join(root, folder, name), os.path.join(root, name)]
    for c in candidates:
        if os.path.exists(c):
            return c, candidates
    return None, candidates


def check_one(binding, root, verbose=True):
    slug = binding['slug']
    sheet = binding['workbook']['sheet']
    lay = binding['layout']
    status = binding.get('status') or {}
    problems, notes = [], []

    path, tried = resolve(binding, root)
    if not path:
        return {'slug': slug, 'verdict': BAD, 'path': None,
                'problems': ['file not found. Tried:\n      ' + '\n      '.join(tried)]}

    # ── the file itself ──────────────────────────────────────────────────────
    a = attrs(path)
    if a is not None:
        if a & FILE_ATTRIBUTE_READONLY:
            problems.append('the file carries the READ-ONLY attribute')
        if a & (FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS | FILE_ATTRIBUTE_RECALL_ON_OPEN):
            problems.append('CLOUD-ONLY placeholder — OneDrive holds no local copy. '
                            'Right-click → Always keep on this device.')
        elif a & FILE_ATTRIBUTE_OFFLINE:
            notes.append('marked offline by OneDrive; first read will download it')

    lock = os.path.join(os.path.dirname(path), '~$' + os.path.basename(path))
    if os.path.exists(lock):
        problems.append('a ~$ lock file exists — somebody has this workbook OPEN')

    # Prove write permission WITHOUT writing: opening r+b takes the write handle
    # and closing it changes nothing — no bytes, no mtime.
    before = os.path.getmtime(path)
    try:
        with open(path, 'r+b'):
            pass
        writable = True
    except PermissionError as e:
        writable = False
        problems.append(f'NOT writable: {e.strerror or e}')
    except OSError as e:
        writable = False
        problems.append(f'could not open for write: {e}')
    if os.path.getmtime(path) != before:
        problems.append('the write probe moved the timestamp — that should be impossible')

    # ── the tab the binding expects ──────────────────────────────────────────
    headers_live = None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            near = [s for s in wb.sheetnames if s.strip().lower() == sheet.strip().lower()]
            problems.append(f'sheet {sheet!r} not found'
                            + (f' (but {near[0]!r} exists — whitespace/case)' if near else ''))
        else:
            ws = wb[sheet]
            a_row, a_col = graph.parse_cell(lay['data_anchor'])
            want = [c['header'] for c in binding['columns']]
            headers_live = []
            for i in range(len(want)):
                v = ws.cell(a_row - 1, a_col + i).value
                headers_live.append('' if v is None else str(v).strip())
            if headers_live != [str(h).strip() for h in want]:
                problems.append(f'header row {a_row - 1} does not match the binding:\n'
                                f'      sheet   {headers_live}\n'
                                f'      binding {want}')
            if ws.cell(a_row, a_col).value in (None, ''):
                notes.append(f'anchor {lay["data_anchor"]} is empty — the tab holds no data today')
        wb.close()
    except Exception as e:                              # noqa: BLE001 - report, never raise
        problems.append(f'could not read the workbook: {type(e).__name__}: {e}')

    # ── the cells we would stamp ─────────────────────────────────────────────
    for key in ('cell', 'block', 'clear'):
        if status.get(key):
            try:
                graph.parse_range(status[key])
            except SystemExit as e:
                problems.append(f'[status] {key} = {status[key]!r} is not a valid range: {e}')

    verdict = BAD if problems else (WARN if notes else OK)
    return {'slug': slug, 'verdict': verdict, 'path': path, 'sheet': sheet,
            'writable': writable, 'problems': problems, 'notes': notes}


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--root', default=CONFIG['local_root'],
                    help='library root, or a folder of test copies')
    ap.add_argument('--slug', action='append', help='check only these bindings')
    args = ap.parse_args()

    root = os.path.expanduser(args.root)
    slugs = args.slug or pivot.list_bindings()
    print(f'root   {root}')
    print(f'checks {len(slugs)} binding(s)')
    print('READ-ONLY: files are opened for writing and closed without writing, '
          'to prove the permission.\n')

    results = []
    for s in slugs:
        b = pivot.load_binding(s)
        r = check_one(b, root)
        results.append(r)
        mark = {OK: '  ok  ', WARN: ' warn ', BAD: ' BAD  '}[r['verdict']]
        where = os.path.basename(r['path']) if r['path'] else '-'
        print(f"[{mark}] {r['slug']:<30} {where:<28} {r.get('sheet', '')}")
        for p in r.get('problems', []):
            print(f'         ! {p}')
        for n in r.get('notes', []):
            print(f'         · {n}')

    bad = [r for r in results if r['verdict'] == BAD]
    warn = [r for r in results if r['verdict'] == WARN]
    print(f"\n{len(results) - len(bad) - len(warn)} ok · {len(warn)} warn · {len(bad)} blocked")
    if bad:
        print('\nBlocked, and each of these would fail at 07:00:')
        for r in bad:
            print(f"  {r['slug']}")
    return 1 if bad else 0


if __name__ == '__main__':
    sys.exit(main())
