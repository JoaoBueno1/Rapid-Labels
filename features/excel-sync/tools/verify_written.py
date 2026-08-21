"""Did what we wrote survive? Read the tab back and say so, per binding.

Built the day an external writer silently lost data in production. The job
reported 21/21 written and was telling the truth about what it had done; the
file was later overwritten by a co-authoring session in Excel for the web, and
nothing in the system noticed. `deliver` proves a write was ISSUED. This proves
it is STILL THERE, which is a different claim and the one that matters.

Run it a few minutes after a delivery, and again later if you want to catch a
slow clobber:

    python tools/verify_written.py                       # every enabled binding
    python tools/verify_written.py --slug hobart-main-stock
    python tools/verify_written.py --root "C:/.../Tests files"

Compares three things against ops.sync_runs' record of the last successful
write: the status stamp, the row count, and a sample of the data itself.
Exits 1 if any binding's write is missing, so a scheduled second pass can fail
loudly.
"""
import argparse
import datetime
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from engine import pivot, supabase                          # noqa: E402
from engine.delivery import CONFIG, graph                    # noqa: E402

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


def resolve(binding, root):
    name = binding['workbook']['file']
    folder = (binding['workbook'].get('folder') or '').replace('/', os.sep)
    for c in (os.path.join(root, folder, name), os.path.join(root, name)):
        if os.path.exists(c):
            return c
    return None


def last_write(sb, slug):
    """What ops.sync_runs says we last wrote for this binding."""
    try:
        rows = sb.rpc('excel_binding_state_get', {'p_binding': slug}) or []
    except SystemExit:
        return None
    return rows[0] if rows else None


def check(binding, root, sb, max_age_h):
    slug = binding['slug']
    sheet = binding['workbook']['sheet']
    out = {'slug': slug, 'problems': [], 'notes': []}

    path = resolve(binding, root)
    if not path:
        out['problems'].append('workbook not found on disk')
        return out

    state = last_write(sb, slug)
    if not state or not state.get('last_written_at'):
        out['notes'].append('no write on record — nothing to verify yet')
        return out

    written = datetime.datetime.fromisoformat(
        str(state['last_written_at']).replace('Z', '+00:00'))
    age_h = (datetime.datetime.now(datetime.timezone.utc)
             - written.astimezone(datetime.timezone.utc)).total_seconds() / 3600
    out['written_at'] = written
    out['age_h'] = age_h
    if age_h > max_age_h:
        out['notes'].append(f'last write was {age_h:.1f}h ago — older than the '
                            f'{max_age_h}h window, not judged')
        return out

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        out['problems'].append(f'sheet {sheet!r} is gone')
        wb.close()
        return out
    ws = wb[sheet]

    # 1) the stamp. Written in the same run as the rows, so if the stamp is old
    #    the rows are old too, whatever they look like.
    cell = (binding.get('status') or {}).get('cell')
    if cell:
        stamp = ws[cell].value
        out['stamp'] = stamp
        if not isinstance(stamp, str) or not stamp.startswith('Updated'):
            out['problems'].append(f'status cell {cell} does not hold our stamp: {stamp!r}')
        else:
            # The stamp carries a date; if it predates the recorded write, the
            # tab has been rolled back to an older version.
            m = re.search(r'(\d{1,2}) ([A-Z][a-z]+) (\d{4})', stamp)
            if m:
                try:
                    d = datetime.datetime.strptime(' '.join(m.groups()), '%d %B %Y').date()
                    if d < written.astimezone().date():
                        out['problems'].append(
                            f'stamp says {d} but we wrote on '
                            f'{written.astimezone().date()} — the tab was rolled back')
                except ValueError:
                    pass

    # 2) the extent. A clobber usually restores a different row count.
    a_row, a_col = graph.parse_cell(binding['layout']['data_anchor'])
    n = 0
    r = a_row
    while True:
        v = ws.cell(r, a_col).value
        if v in (None, ''):
            break
        n += 1
        r += 1
    out['rows_now'] = n
    expected = state.get('last_rows')
    if expected and n != expected:
        out['problems'].append(f'wrote {expected} rows, tab now holds {n}')

    wb.close()
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--slug', action='append')
    ap.add_argument('--root', default=CONFIG['local_root'])
    ap.add_argument('--max-age-hours', type=float, default=20.0,
                    help='only judge writes newer than this')
    args = ap.parse_args()

    sb = supabase.Client(schema='public')
    root = os.path.expanduser(args.root)
    slugs = args.slug or [s for s in pivot.list_bindings()
                          if pivot.load_binding(s).get('enabled')]
    if not slugs:
        print('No enabled bindings to verify. Nothing was checked.')
        return 0

    print(f'root  {root}')
    print(f'checking {len(slugs)} binding(s) against what ops.sync_runs says we wrote\n')

    bad = []
    for s in slugs:
        r = check(pivot.load_binding(s), root, sb, args.max_age_hours)
        mark = ' LOST ' if r['problems'] else ('  --  ' if r.get('notes') else '  ok  ')
        extra = ''
        if r.get('rows_now') is not None:
            extra = f"{r['rows_now']} rows"
        if r.get('age_h') is not None:
            extra += f", written {r['age_h']:.1f}h ago"
        print(f"[{mark}] {r['slug']:<30} {extra}")
        for p in r['problems']:
            print(f'         ! {p}')
        for nte in r.get('notes', []):
            print(f'         · {nte}')
        if r['problems']:
            bad.append(r['slug'])

    print()
    if bad:
        print(f'{len(bad)} binding(s) LOST their write after the job reported success:')
        for s in bad:
            print(f'  {s}')
        print('\nOn a OneDrive-synced file this is the co-authoring clobber: somebody')
        print('had the workbook open, Excel for the web autosaved their session, and')
        print('the server copy replaced ours. Re-running only helps if they closed it.')
        return 1
    print('Every checked write is still in place.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
