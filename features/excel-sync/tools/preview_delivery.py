"""What would change in each tab if we switched delivery on? Read-only.

A dry run proves the write is *safe* — the header still matches, no formula is
in the way, the row count is sane. It says nothing about whether the numbers
are *right*, and those are different questions. This one answers the second:
for every binding, it diffs the grid the engine would write against what the tab
holds today, key by key.

That is the review nobody wants to do by opening seven workbooks. It is also the
only way to see the thing worth seeing before go-live — a tab that has quietly
drifted. Brisbane's `Sales MTD` currently holds June figures under a hand-typed
"Updated 10-Aug"; against an August dataset that shows up here as a near-total
rewrite, which is the correct and alarming answer.

    python tools/preview_delivery.py                     # every binding
    python tools/preview_delivery.py brisbane-main-stock
    python tools/preview_delivery.py --root out          # against test copies
    python tools/preview_delivery.py --samples 15

Uses openpyxl, not COM: it never opens Excel and never touches the file beyond
reading it, so it is safe to run against the live library while people work.
"""
import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from engine import flat, pivot, supabase                    # noqa: E402
from engine.delivery import CONFIG, graph                    # noqa: E402

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


def _n(v):
    """Compare 4 and '4' and 4.0 as equal — the tabs were pasted from a report
    and carry numbers as text in places, which is not a difference anyone means."""
    if v is None or v == '':
        return ''
    if isinstance(v, bool):
        return str(v)
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else f'{f:.4f}'
    except (TypeError, ValueError):
        return str(v).strip()


def read_tab(path, sheet, anchor, n_cols):
    """The rectangle the binding owns, as it stands today: {key: [values]}."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit(f'sheet {sheet!r} not in {os.path.basename(path)}')
    ws = wb[sheet]
    a_row, a_col = graph.parse_cell(anchor)
    out, order = {}, []
    for row in ws.iter_rows(min_row=a_row, min_col=a_col,
                            max_col=a_col + n_cols - 1, values_only=True):
        key = row[0]
        if key in (None, ''):
            break                       # the block ends at the first blank key
        k = str(key).strip()
        out[k] = [_n(v) for v in row]
        order.append(k)
    wb.close()
    return out, order


def preview(slug, sb, root_override=None, samples=8):
    b = pivot.load_binding(slug)
    wb_spec = b['workbook']
    folder = (wb_spec.get('folder') or '').replace('/', os.sep)
    if root_override:
        path = os.path.join(root_override, wb_spec['file'])
    else:
        path = os.path.join(CONFIG['local_root'], folder, wb_spec['file'])
    if not os.path.exists(path):
        print(f'  {slug:<30} SKIP — not on disk: {path}')
        return None

    rows = flat.fetch_dataset(b['dataset'], sb)
    tab = flat.build(rows, b)
    cols = tab['cols']
    new = {r[0].strip() if isinstance(r[0], str) else str(r[0]): [_n(v) for v in r]
           for r in tab['grid']}

    cur, _ = read_tab(path, wb_spec['sheet'], b['layout']['data_anchor'], len(cols))

    added = [k for k in new if k not in cur]
    removed = [k for k in cur if k not in new]
    both = [k for k in new if k in cur]
    changed = [k for k in both if new[k] != cur[k]]

    same_pct = 100.0 * (len(both) - len(changed)) / max(len(new), 1)
    flag = '' if same_pct >= 80 else '   <<< large rewrite, look at it'
    print(f'\n  {slug}  ({wb_spec["sheet"]})')
    print(f'    tab {len(cur)} rows -> dataset {len(new)} rows   '
          f'| unchanged {len(both) - len(changed)} ({same_pct:.0f}%) '
          f'| changed {len(changed)} | new {len(added)} | dropped {len(removed)}{flag}')

    headers = [c['header'] for c in cols]
    for k in changed[:samples]:
        diffs = [f'{headers[i]} {cur[k][i]!r}->{new[k][i]!r}'
                 for i in range(len(cols)) if cur[k][i] != new[k][i]]
        print(f'      ~ {k:<28} ' + '; '.join(diffs[:4]))
    if len(changed) > samples:
        print(f'      … and {len(changed) - samples} more changed')
    for k in added[:3]:
        print(f'      + {k}')
    if len(added) > 3:
        print(f'      … and {len(added) - 3} more new')
    for k in removed[:3]:
        print(f'      - {k}   (in the tab today, not in the dataset)')
    if len(removed) > 3:
        print(f'      … and {len(removed) - 3} more dropped')

    return {'slug': slug, 'tab_rows': len(cur), 'new_rows': len(new),
            'changed': len(changed), 'added': len(added), 'removed': len(removed),
            'same_pct': same_pct}


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('slug', nargs='*', help='binding slugs (default: all)')
    ap.add_argument('--root', help='look for workbooks here instead of the synced library')
    ap.add_argument('--samples', type=int, default=8)
    args = ap.parse_args()

    sb = supabase.Client(schema='public')
    slugs = args.slug or pivot.list_bindings()
    print(f'{len(slugs)} binding(s). READ-ONLY — opens the workbooks with openpyxl, '
          'writes nothing.')

    results = []
    for s in slugs:
        try:
            r = preview(s, sb, args.root, args.samples)
            if r:
                results.append(r)
        except SystemExit as e:
            print(f'  {s:<30} ERROR — {e}')

    if results:
        print('\n' + '=' * 96)
        print(f"{'binding':<32}{'tab':>7}{'new':>7}{'chg':>7}{'add':>6}{'drop':>6}{'same':>8}")
        print('-' * 96)
        for r in sorted(results, key=lambda x: x['same_pct']):
            print(f"{r['slug']:<32}{r['tab_rows']:>7}{r['new_rows']:>7}"
                  f"{r['changed']:>7}{r['added']:>6}{r['removed']:>6}{r['same_pct']:>7.0f}%")
    return 0


if __name__ == '__main__':
    sys.exit(main())
