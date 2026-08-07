"""Spec loading, the SKU x warehouse pivot, and local delivery (csv / xlsx).

The pivot shape is Cin7's: one row of merged warehouse names above one row of
metric names, then the data. Columns are matched BY HEADER NAME everywhere —
never by index. Cin7 has re-ordered report columns before (the lesson baked
into scripts/update_main_avg_3mo.py), and a spec that trusts position turns
that into silently shifted data.
"""
import csv
import os
import tomllib

SPECS = os.path.join(os.path.dirname(__file__), '..', 'specs')
DATASET_DIR = os.path.join(SPECS, 'datasets')
BINDING_DIR = os.path.join(SPECS, 'bindings')

# Datasets and bindings are separate on purpose. Several company workbooks want
# the same numbers — stock availability, monthly sales — in different tabs,
# column subsets and cadences. One dataset is built and stored once; each
# workbook is a binding that reads it. Connecting a new spreadsheet is then a
# file in specs/bindings/, with no new query, sync or Cin7 call.


def _load(path, slug, required):
    if not os.path.exists(path):
        raise SystemExit(f'No spec at {path}')
    with open(path, 'rb') as fh:
        spec = tomllib.load(fh)
    for need in required:
        if need not in spec:
            raise SystemExit(f'{os.path.basename(path)} is missing [{need}]')
    if spec['slug'] != slug:
        raise SystemExit(f"{os.path.basename(path)} declares slug='{spec['slug']}'")
    return spec


def load_spec(slug):
    return _load(os.path.join(DATASET_DIR, f'{slug}.toml'), slug,
                 ('slug', 'source', 'layout', 'metrics'))


def load_binding(slug):
    return _load(os.path.join(BINDING_DIR, f'{slug}.toml'), slug,
                 ('slug', 'dataset', 'workbook'))


def list_specs():
    return sorted(f[:-5] for f in os.listdir(DATASET_DIR) if f.endswith('.toml'))


def list_bindings():
    if not os.path.isdir(BINDING_DIR):
        return []
    return sorted(f[:-5] for f in os.listdir(BINDING_DIR) if f.endswith('.toml'))


def binding_view(spec, binding):
    """A binding may take a subset of the dataset's columns, in its own order.
    Returns a spec-shaped dict the pivot can build from unchanged."""
    wanted = binding.get('columns')
    if not wanted:
        return spec
    by_header = {m['header']: m for m in spec['metrics']}
    missing = [h for h in wanted if h not in by_header]
    if missing:
        raise SystemExit(f"binding '{binding['slug']}' asks for columns the "
                         f"'{spec['slug']}' dataset does not produce: {missing}")
    view = dict(spec)
    view['metrics'] = [by_header[h] for h in wanted]
    return view


def build(rows, spec):
    """Long rows -> {groups, keys, cells, grid} ready for delivery."""
    layout = spec['layout']
    key_field = layout.get('key', 'sku')
    group_field = layout.get('group', 'location')
    fixed = spec.get('fixed', [])
    metrics = spec['metrics']

    cells = {}
    for r in rows:
        cells[(r[key_field], r[group_field])] = r

    groups = sorted({r[group_field] for r in rows})
    if not layout.get('drop_empty_groups', True) and layout.get('all_groups'):
        groups = sorted(set(groups) | set(layout['all_groups']))
    # Cin7 sorts SKUs case-insensitively; match it so a human eyeballing the two
    # side by side sees the same order.
    keys = sorted({r[key_field] for r in rows}, key=lambda s: (str(s).lower(), str(s)))

    # first row of a key wins for the fixed columns (Unit is per product)
    first = {}
    for r in rows:
        first.setdefault(r[key_field], r)

    group_row = [''] * len(fixed)
    header_row = [f['header'] for f in fixed]
    for g in groups:
        for m in metrics:
            group_row.append(g)
            header_row.append(m['header'])

    grid = []
    for k in keys:
        line = [first[k].get(f['field'], '') for f in fixed]
        for g in groups:
            cell = cells.get((k, g))
            for m in metrics:
                if cell is None:
                    line.append('')          # absent block stays BLANK, not 0
                    continue
                v = cell.get(m['field'])
                if v is None:
                    line.append('')
                elif isinstance(v, float):
                    nd = m.get('round')
                    line.append(round(v, nd) if nd is not None else v)
                else:
                    line.append(v)
        grid.append(line)

    return {'groups': groups, 'keys': keys, 'cells': cells,
            'group_row': group_row, 'header_row': header_row, 'grid': grid}


def preamble_rows(spec, meta):
    """Cin7 puts 4 metadata lines above the header. Reproduced so a local build
    can be eyeballed against the export it is meant to replace."""
    if not spec['layout'].get('preamble'):
        return []
    period = meta.get('period', '')
    frm, to = (period.split('..') + ['', ''])[:2] if period else ('', '')
    return [[f"Report period: {spec.get('period_label', 'This month')}"],
            [f'From: {frm}'], [f'To: {to}'], ['Currency: Base Currency']]


def write_csv(built, spec, meta, path):
    with open(path, 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh)
        for r in preamble_rows(spec, meta):
            w.writerow(r)
        w.writerow(built['group_row'])
        w.writerow(built['header_row'])
        w.writerows(built['grid'])
    return path


def write_xlsx(built, spec, meta, path):
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec.get('workbook', {}).get('sheet') or 'Sheet'
    for r in preamble_rows(spec, meta):
        ws.append(r)
    ws.append(built['group_row'])
    ws.append(built['header_row'])
    bold = Font(bold=True)
    for row in (ws.max_row - 1, ws.max_row):
        for c in ws[row]:
            c.font = bold
    for line in built['grid']:
        ws.append(line)
    ws.freeze_panes = ws.cell(row=ws.max_row - len(built['grid']) + 1, column=len(spec.get('fixed', [])) + 1)
    wb.save(path)
    return path
