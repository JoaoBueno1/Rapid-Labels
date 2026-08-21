"""Write dataset blocks straight into a workbook on disk.

Why this exists alongside the Graph adapter: the branch workbooks live in a
SharePoint library that is synced to the operator's Mac, so writing the local
file IS writing to SharePoint — the sync client uploads it. That skips Microsoft
Graph, the Entra app and the admin consent entirely.

Trade-off, stated plainly: it needs that machine awake with OneDrive running.
Less robust than Graph on a schedule, but available today and it reuses the same
datasets, specs and bindings — nothing is thrown away when Graph lands.

What a round-trip through openpyxl costs, measured on the real Coffs Harbour
workbook (2.20 MB, 40 sheets):

    sheets, formulas, conditional formatting, merged cells   preserved
    drawings/, media/, customXml/                            LOST (32 parts)
    calcChain.xml                                            dropped; Excel rebuilds it

So this is safe for these workbooks only because the images sit outside the data
area and are being retired anyway. `assert_no_drawings()` fails loudly if a
target ever gains one, rather than deleting it quietly on the next run.
"""
import os
import re
import shutil
import zipfile
from datetime import datetime

CLOUD_MARKERS = ('/Library/CloudStorage/', '/OneDrive', '/SharePoint')


def is_synced_path(path):
    """True when the file lives in a cloud-synced folder — i.e. writing it
    publishes to everyone. Used to force an explicit opt-in."""
    p = os.path.abspath(path)
    return any(m in p for m in CLOUD_MARKERS)


def assert_writable(path, allow_synced=False):
    if is_synced_path(path) and not allow_synced:
        raise SystemExit(
            f'REFUSING to write {path}\n'
            '  That path is inside a cloud-synced folder, so the change would reach\n'
            '  everyone using the file. Pass allow_synced=True (CLI: --i-know-this-is-live)\n'
            '  only when that is genuinely intended.')


def inspect_parts(path):
    """What is in the file today, so a loss can be proven rather than assumed."""
    names = zipfile.ZipFile(path).namelist()
    return {
        'drawings': [n for n in names if n.startswith('xl/drawings/')],
        'media': [n for n in names if n.startswith('xl/media/')],
        'customXml': [n for n in names if n.startswith('customXml/')],
        'total': len(names),
    }


def assert_no_drawings(path, sheets):
    """Refuse to run if a target sheet has a drawing openpyxl would silently drop."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    bad = []
    for s in sheets:
        if s not in wb.sheetnames:
            raise SystemExit(f'{os.path.basename(path)}: no sheet named {s!r}')
        ws = wb[s]
        n = len(getattr(ws, '_images', [])) + len(getattr(ws, '_charts', []))
        if n:
            bad.append(f'{s} ({n})')
    wb.close()
    if bad:
        raise SystemExit(
            'target sheets carry drawings/charts that this writer would destroy: '
            + ', '.join(bad) + '\n  Move them off the data sheets, or use the Graph adapter.')


def _cell_ref(anchor):
    m = re.match(r'^([A-Z]+)(\d+)$', anchor.upper())
    if not m:
        raise SystemExit(f'bad anchor {anchor!r}')
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return col, int(m.group(2))


def write_blocks(path, jobs, backup_dir=None, allow_synced=False, dry_run=False):
    """Apply one or more blocks to a workbook.

    Each job: {sheet, anchor, grid, status_cell?, status_text?}

    Rows below the new block are CLEARED. Without that, a month where the SKU
    count shrinks would leave yesterday's rows alive at the bottom, silently
    mixing two days of data — the failure mode nobody notices.
    """
    import openpyxl
    assert_writable(path, allow_synced)
    assert_no_drawings(path, [j['sheet'] for j in jobs])

    report = {'file': path, 'sheets': [], 'backup': None}

    if backup_dir and not dry_run:
        os.makedirs(backup_dir, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        dest = os.path.join(backup_dir, f'{stamp}--{os.path.basename(path)}')
        shutil.copy2(path, dest)
        report['backup'] = dest

    wb = openpyxl.load_workbook(path)
    for job in jobs:
        ws = wb[job['sheet']]
        c0, r0 = _cell_ref(job['anchor'])
        grid = job['grid']
        width = max((len(r) for r in grid), default=0)

        # how far the previous block reached, by the key column
        prev_last = r0 - 1
        r = r0
        while r <= ws.max_row and ws.cell(r, c0).value not in (None, ''):
            prev_last = r
            r += 1

        if not dry_run:
            for i, row in enumerate(grid):
                for j, v in enumerate(row):
                    ws.cell(r0 + i, c0 + j).value = v
            new_last = r0 + len(grid) - 1
            cleared = 0
            for rr in range(new_last + 1, prev_last + 1):
                for j in range(width):
                    ws.cell(rr, c0 + j).value = None
                cleared += 1
            if job.get('status_cell'):
                ws[job['status_cell']] = job.get('status_text', '')
        else:
            new_last = r0 + len(grid) - 1
            cleared = max(0, prev_last - new_last)

        report['sheets'].append({
            'sheet': job['sheet'], 'anchor': job['anchor'],
            'rows': len(grid), 'cols': width,
            'prev_last_row': prev_last, 'new_last_row': new_last,
            'rows_cleared': cleared,
            'status_cell': job.get('status_cell'),
        })

    if not dry_run:
        wb.save(path)
    wb.close()
    return report


def status_text(built_at=None, next_run=None, rows=None):
    """The 'Updated: 03-Aug' cell the tabs already carry, written for real.
    Kept short — it shares row 1 with the group headers."""
    now = built_at or datetime.now()
    parts = [f'Updated: {now.strftime("%d-%b %H:%M")}']
    if rows is not None:
        parts.append(f'{rows:,} rows')
    if next_run:
        parts.append(f'next {next_run}')
    return ' · '.join(parts)
