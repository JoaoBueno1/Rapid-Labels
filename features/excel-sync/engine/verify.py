"""Gates that must pass before a build is allowed out, and the diff that proves
a build reproduces the Cin7 report it replaces.

The gates exist because of one specific failure mode: a report that is 6% low
looks completely normal in a spreadsheet, while a tab that did not refresh is
obvious. Better to block the write than to publish a quiet wrong number.
"""


def check(meta, spec, built):
    """Return (ok, [(level, message)]). level is 'block' or 'warn'."""
    v = spec.get('validate', {})
    out = []

    need = v.get('min_detail_coverage_pct')
    if need is not None and 'detail_coverage_pct' in meta:
        got = meta['detail_coverage_pct']
        lvl = 'block' if got < need else 'ok'
        out.append((lvl, f'order detail coverage {got:.2f}% (need >= {need}%)'
                    f' — {meta["orders_with_detail"]}/{meta["orders_counted"]} orders'))

    lo_hi = v.get('expect_rows_between')
    if lo_hi:
        n = len(built['keys'])
        lvl = 'ok' if lo_hi[0] <= n <= lo_hi[1] else 'block'
        out.append((lvl, f'{n} rows (expected {lo_hi[0]}–{lo_hi[1]})'))

    mg = v.get('min_groups')
    if mg is not None:
        n = len(built['groups'])
        out.append(('ok' if n >= mg else 'block', f'{n} warehouse columns (need >= {mg})'))

    if not built['grid']:
        out.append(('block', 'empty result set'))

    ok = not any(l == 'block' for l, _ in out)
    return ok, out


# ─────────────────────────────────────────────────────────────────────────────
def read_cin7_export(path, sheet=None):
    """Parse a Cin7 export into {(sku, warehouse): {metric: value}}.

    Handles both shapes seen so far: a 2-row header at the top (stock level) and
    one preceded by 4 metadata lines (sale order details). The header row is
    found by looking for the cell that says SKU, not by a fixed offset.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    hdr_i = None
    for i, r in enumerate(rows[:12]):
        if r and str(r[0]).strip() == 'SKU':
            hdr_i = i
            break
    if hdr_i is None:
        raise SystemExit(f'{path}: no header row starting with "SKU" in the first 12 rows')

    header = [str(c).strip() if c is not None else '' for c in rows[hdr_i]]
    group_raw = [str(c).strip() if c is not None else '' for c in rows[hdr_i - 1]] if hdr_i else []
    group_raw += [''] * (len(header) - len(group_raw))

    cur = ''
    groups = []
    for g in group_raw:                    # merged cells arrive as blanks
        if g:
            cur = g
        groups.append(cur)

    meta = {}
    for r in rows[:hdr_i - 1 if hdr_i else 0]:
        first = str(r[0]).strip() if r and r[0] is not None else ''
        if ':' in first:
            k, _, val = first.partition(':')
            meta[k.strip()] = val.strip()

    cells = {}
    for r in rows[hdr_i + 1:]:
        if not r or r[0] in (None, ''):
            continue
        sku = str(r[0]).strip()
        by_group = {}
        for i in range(1, len(header)):
            if not header[i] or not groups[i]:
                continue
            by_group.setdefault(groups[i], {})[header[i]] = r[i] if i < len(r) else None
        for g, vals in by_group.items():
            if all(v in (None, '') for v in vals.values()):
                continue                    # block absent for this SKU
            cells[(sku, g)] = vals
    return {'meta': meta, 'cells': cells, 'groups': sorted(set(g for g in groups if g))}


def diff(built, spec, export, key_metric=None):
    """Compare a build against a parsed Cin7 export, per metric."""
    metrics = spec['metrics']
    key_metric = key_metric or metrics[0]['header']
    exp = export['cells']
    ours = {}
    for (k, g), row in built['cells'].items():
        ours[(k, g)] = row

    both = [c for c in exp if c in ours]
    only_exp = [c for c in exp if c not in ours]
    only_mir = [c for c in ours if c not in exp]

    per_metric = {}
    for m in metrics:
        same = 0
        diffs = []
        tol = m.get('tolerance', 0.005)
        for c in both:
            e = exp[c].get(m['header'])
            o = ours[c].get(m['field'])
            e = 0.0 if e in (None, '') else float(e)
            o = 0.0 if o in (None, '') else float(o)
            if abs(e - o) <= max(tol, abs(e) * 1e-6):
                same += 1
            else:
                diffs.append((c, e, o, o - e))
        tot_e = sum(float(exp[c].get(m['header']) or 0) for c in exp)
        tot_o = sum(float(ours[c].get(m['field']) or 0) for c in ours)
        diffs.sort(key=lambda d: -abs(d[3]))
        per_metric[m['header']] = {
            'same': same, 'both': len(both), 'diffs': diffs,
            'total_export': tot_e, 'total_mirror': tot_o,
            'pct': (100.0 * tot_o / tot_e) if tot_e else 0.0,
        }
    return {'both': len(both), 'only_export': only_exp, 'only_mirror': only_mir,
            'metrics': per_metric, 'export_cells': len(exp), 'mirror_cells': len(ours)}
