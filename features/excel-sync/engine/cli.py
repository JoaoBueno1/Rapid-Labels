"""excel-sync CLI.

    python -m engine list
    python -m engine build monthly-sales [--month 2026-08] [--out DIR]
    python -m engine build monthly-sales --verify "~/Downloads/Sale Order Details (23).xlsx"

`--verify` is the point of phase 1: build the tab from the mirror and diff it,
cell by cell, against the Cin7 export it is meant to replace. Nothing writes to
a real workbook until that diff is clean — Microsoft Graph is a later phase and
a swappable adapter, not part of this path.
"""
import argparse
import os
import sys
from datetime import datetime

from . import flat, pivot, sources, verify

# A Windows console defaults to cp1252, which cannot encode the arrows and
# bullets this output uses — `list` died on a UnicodeEncodeError before printing
# its bindings. CI is UTF-8 already; this is for the machine the first real
# write will be run from.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


def _fmt(v):
    return f'{round(v):,}'.replace(',', ' ') if isinstance(v, (int, float)) else str(v)


def cmd_list(_):
    print('datasets (built once, stored in excel_sync.dataset_rows):')
    for slug in pivot.list_specs():
        spec = pivot.load_spec(slug)
        cols = ', '.join(m['header'] for m in spec['metrics'])
        print(f"  {slug:<16} {spec.get('title', '')}")
        print(f"  {'':<16} source={spec['source']}  columns: {cols}")
    bindings = pivot.list_bindings()
    print(f'\nbindings (workbook tabs consuming them): {len(bindings) or "none yet"}')
    for slug in bindings:
        b = pivot.load_binding(slug)
        wb = b['workbook']
        state = 'enabled' if b.get('enabled') else 'disabled'
        print(f"  {slug:<16} {b['dataset']} → {wb.get('file')} / {wb.get('sheet')}  [{state}]")
    if not bindings:
        print('  add one: see specs/bindings/README.md')
    return 0


def cmd_register(_):
    """Mirror datasets and bindings into ops.sync_registry so the monitor lists them."""
    from . import publish
    specs = [pivot.load_spec(s) for s in pivot.list_specs()]
    publish.register_datasets(specs)
    print(f'  registered {len(specs)} dataset build(s)')
    for s in specs:
        print(f"    excel-dataset-{s['slug']}")

    bindings = [pivot.load_binding(s) for s in pivot.list_bindings()]
    if bindings:
        publish.register_bindings(bindings)
        print(f'  registered {len(bindings)} workbook binding(s)')
        for b in bindings:
            print(f"    excel-{b['slug']:<20} {b['dataset']} → {b['workbook'].get('file')}")
    else:
        print('  no workbook bindings yet — see specs/bindings/README.md')
    return 0


def cmd_build(args):
    spec = pivot.load_spec(args.slug)
    fn = sources.SOURCES.get(spec['source'])
    if fn is None:
        raise SystemExit(f"spec references unknown source '{spec['source']}'")

    print(f"▶ {args.slug} — {spec.get('title', '')}")
    rows, meta = fn(month=args.month)
    built = pivot.build(rows, spec)
    print(f"  source .......... {meta.get('cells')} cells, {meta.get('calls')} Supabase calls")
    if 'period' in meta:
        print(f"  period .......... {meta['period']}")
    print(f"  pivot ........... {len(built['keys'])} rows x {len(built['groups'])} warehouses")

    ok, checks = verify.check(meta, spec, built)
    print('  gates:')
    for level, msg in checks:
        mark = {'ok': '✓', 'warn': '!', 'block': '✗'}[level]
        print(f'    {mark} {msg}')

    if args.verify:
        path = os.path.expanduser(args.verify)
        if not os.path.exists(path):
            raise SystemExit(f'export not found: {path}')
        export = verify.read_cin7_export(path, spec.get('workbook', {}).get('source_sheet'))
        d = verify.diff(built, spec, export)
        print(f"\n  ── diff vs {os.path.basename(path)}")
        if export['meta']:
            print('     export says:', ' | '.join(f'{k}: {v}' for k, v in export['meta'].items()))
        print(f"     cells  export {d['export_cells']}   mirror {d['mirror_cells']}   "
              f"in both {d['both']}   only-export {len(d['only_export'])}   only-mirror {len(d['only_mirror'])}")
        # Tolerance is per metric: Quantity and Total fail for different reasons
        # and at different thresholds, so one global number would either hide a
        # money error or block on a known quantity quirk.
        by_header = {m['header']: m for m in spec['metrics']}
        failed = []
        for name, r in d['metrics'].items():
            pct = 100.0 * r['same'] / r['both'] if r['both'] else 0.0
            off = abs(100.0 - r['pct'])
            tol = by_header[name].get('total_tolerance_pct', args.tolerance)
            mark = '✓' if off <= tol else '✗'
            if off > tol:
                failed.append((name, off, tol))
            print(f"     {mark} {name:<16} identical {r['same']}/{r['both']} ({pct:.2f}%)   "
                  f"total {_fmt(r['total_mirror'])} vs {_fmt(r['total_export'])} = {r['pct']:.2f}%  (tol ±{tol}%)")
            for (k, g), e, o, delta in r['diffs'][:args.show]:
                print(f"        {k} | {g}: export={_fmt(e)} mirror={_fmt(o)} Δ={'+' if delta > 0 else ''}{_fmt(delta)}")
        for c in d['only_export'][:args.show]:
            print(f"     only in export: {c[0]} | {c[1]}")
        for c in d['only_mirror'][:args.show]:
            print(f"     only in mirror: {c[0]} | {c[1]}")
        if failed:
            for name, off, tol in failed:
                print(f'\n  ✗ {name} off by {off:.2f}% — above its {tol}% tolerance')
            ok = False
        else:
            print('\n  ✓ every metric within tolerance of the Cin7 export')

    if args.out:
        os.makedirs(args.out, exist_ok=True)
        stem = args.slug + (f"-{meta['month']}" if meta.get('month') else '')
        csv_path = pivot.write_csv(built, spec, meta, os.path.join(args.out, stem + '.csv'))
        print(f'\n  wrote {csv_path}')
        xl = pivot.write_xlsx(built, spec, meta, os.path.join(args.out, stem + '.xlsx'))
        print(f'  wrote {xl}' if xl else '  (openpyxl missing — csv only)')

    if args.publish:
        from . import publish as pub
        with pub.Run('excel-dataset-' + args.slug) as run:
            if not ok and not args.force:
                run.finish('blocked', 'validation gates failed')
                print('\n  BLOCKED — nothing published')
                return 1
            digest, changed, written = pub.publish_dataset(spec, rows, meta, built)
            run.rows_written = len(rows)
            run.stats = {'checksum': digest, 'changed': changed,
                         'groups': len(built['groups']), 'keys': len(built['keys'])}
            run.finish('success')
        print(f"\n  published {len(rows)} rows to excel_sync.dataset_rows"
              f"  checksum={digest[:12]}  {'CHANGED' if changed else 'unchanged since last build'}")

    if not ok:
        print('\n  BLOCKED — not fit to publish')
        return 1
    print('\n  OK')
    return 0


def cmd_render(args):
    """Build the exact block a workbook tab will receive, and diff it against the
    tab as it stands today. This is the binding-level equivalent of `--verify`:
    proof before anything is ever written through Graph."""
    binding = pivot.load_binding(args.slug)
    lay = binding['layout']
    print(f"▶ {args.slug} — {binding.get('title','')}")
    print(f"  {binding['dataset']}  →  {binding['workbook']['file']} / {binding['workbook']['sheet']}")
    print(f"  locations: {', '.join(lay['locations'])}"
          + ("  (SUMMED)" if len(lay['locations']) > 1 else ""))

    rows = flat.fetch_dataset(binding['dataset'])
    built = flat.build(rows, binding)
    rng = flat.a1(lay['data_anchor'], len(built['grid']), len(built['cols']))
    print(f"  block ......... {len(built['grid'])} rows x {len(built['cols'])} cols  →  {rng}")
    if built['missing_locations']:
        print(f"  ! locations with no rows in the dataset: {built['missing_locations']}")

    if args.workbook:
        import openpyxl
        wb = openpyxl.load_workbook(os.path.expanduser(args.workbook), data_only=True)
        sheet = binding['workbook']['sheet']
        if sheet not in wb.sheetnames:
            raise SystemExit(f"workbook has no sheet {sheet!r}")
        ws = wb[sheet]
        import re
        m = re.match(r'^([A-Z]+)(\d+)$', lay['data_anchor'].upper())
        c0 = 0
        for ch in m.group(1):
            c0 = c0 * 26 + (ord(ch) - 64)
        r0 = int(m.group(2))
        cur = {}
        r = r0
        while r <= ws.max_row:
            k = ws.cell(r, c0).value
            if k in (None, ''):
                break
            cur[str(k).strip()] = [ws.cell(r, c0 + i).value for i in range(1, len(built['cols']))]
            r += 1
        print(f"\n  ── tab today: {len(cur)} rows   vs   ours: {len(built['grid'])}")
        ours = {row[0]: row[1:] for row in built['grid']}
        both = [k for k in cur if k in ours]
        print(f"     SKUs in both {len(both)} | only in tab {len(set(cur) - set(ours))}"
              f" | only in ours {len(set(ours) - set(cur))}")
        # compare the column every formula actually reads
        idx = next((i for i, c in enumerate(built['cols'][1:])
                    if c['field'] in ('available', 'quantity')), None)
        if idx is not None and both:
            name = built['cols'][idx + 1]['header']
            same = 0
            diffs = []
            for k in both:
                a = cur[k][idx]
                b = ours[k][idx]
                a = 0 if a in (None, '') else float(a)
                b = 0 if b in (None, '') else float(b)
                if abs(a - b) < 0.005:
                    same += 1
                else:
                    diffs.append((k, a, b))
            print(f"     '{name}' identical on {same}/{len(both)} ({100*same/len(both):.1f}%)"
                  f"  — the column the VLOOKUPs read")
            diffs.sort(key=lambda d: -abs(d[2] - d[1]))
            for k, a, b in diffs[:args.show]:
                print(f"        {k:<28} tab={a:>10.0f}  ours={b:>10.0f}  Δ={b-a:+.0f}")

    if args.out:
        os.makedirs(args.out, exist_ok=True)
        import csv
        path = os.path.join(args.out, args.slug + '.csv')
        with open(path, 'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow([c['header'] for c in built['cols']])
            w.writerows(built['grid'])
        print(f"\n  wrote {path}")
    return 0


def cmd_graph_login(_):
    """One browser sign-in; after this the nightly run renews its own token."""
    from .delivery import auth
    auth.device_login()
    return 0


def cmd_deliver(args):
    """Write bindings into their workbook tabs. Dry run unless --write.

    Runs are logged to ops.sync_runs only when something is actually written —
    a dry run is a rehearsal and has no business showing up on the monitor as a
    sync that happened.
    """
    from . import delivery, publish as pub

    slugs = [args.slug] if args.slug else pivot.list_bindings()
    rc = 0
    for slug in slugs:
        binding = pivot.load_binding(slug)
        if not args.write:
            res = delivery.deliver(binding, write=False, force=args.force,
                                   mode=args.mode, file_override=args.file,
                                   transport=args.transport, root=args.root)
            rc = rc or (1 if res.get('status') == 'blocked' else 0)
            continue

        with pub.Run('excel-' + slug) as run:
            res = delivery.deliver(binding, write=True, force=args.force,
                                   mode=args.mode, file_override=args.file,
                                   transport=args.transport, root=args.root)
            run.rows_written = res.get('rows')
            run.stats = {k: v for k, v in res.items()
                         if k in ('address', 'cols', 'batches', 'clear',
                                  'formula_cells_on_tab', 'file', 'sheet')}
            if res.get('wrote'):
                run.finish('success')
            elif res.get('status') == 'disabled':
                run.finish('skipped', 'binding disabled')
            else:
                run.finish('blocked', res.get('error', res.get('status')))
                rc = 1
    return rc
def cmd_apply(args):
    """Build every binding for a workbook and write them in one pass."""
    from .delivery import local_xlsx
    path = os.path.expanduser(args.workbook)
    if not os.path.exists(path):
        raise SystemExit(f'not found: {path}')

    parts = local_xlsx.inspect_parts(path)
    print(f"▶ {os.path.basename(path)}")
    print(f"  parts: {parts['total']}  drawings={len(parts['drawings'])} "
          f"media={len(parts['media'])} customXml={len(parts['customXml'])}"
          + ("   ← these are LOST on write" if (parts['drawings'] or parts['media']) else ""))
    if local_xlsx.is_synced_path(path):
        print("  ⚠ this path is CLOUD-SYNCED — writing it publishes to everyone")

    jobs = []
    for slug in args.bindings:
        binding = pivot.load_binding(slug)
        lay = binding['layout']
        rows = flat.fetch_dataset(binding['dataset'])
        built = flat.build(rows, binding)
        rng = flat.a1(lay['data_anchor'], len(built['grid']), len(built['cols']))
        print(f"  {slug:<18} {binding['workbook']['sheet']:<11} {len(built['grid']):>5} rows → {rng}"
              + ("  (SUMMED)" if len(lay['locations']) > 1 else ""))
        jobs.append({
            'sheet': binding['workbook']['sheet'],
            'anchor': lay['data_anchor'],
            'grid': built['grid'],
            'status_cell': binding.get('status', {}).get('cell'),
            'status_text': local_xlsx.status_text(rows=len(built['grid'])),
        })

    rep = local_xlsx.write_blocks(
        path, jobs, backup_dir=args.backup_dir,
        allow_synced=args.i_know_this_is_live, dry_run=args.dry_run)

    print(f"\n  {'would write' if args.dry_run else 'wrote'}:")
    for s in rep['sheets']:
        note = f"  cleared {s['rows_cleared']} leftover rows" if s['rows_cleared'] else ""
        print(f"    {s['sheet']:<11} {s['rows']} rows at {s['anchor']} "
              f"(last row {s['prev_last_row']} → {s['new_last_row']}){note}")
        if s['status_cell']:
            print(f"    {'':11} status → {s['status_cell']}")
    if rep['backup']:
        print(f"  backup: {rep['backup']}")
    return 0


def cmd_master(args):
    """Build one workbook holding every dataset, one tab each.

    Not part of the delivery path — the branch workbooks are written directly,
    because linking them to an external master would mean rewiring 6,600 internal
    formulas AND relying on cross-file links, which barely refresh in Excel for
    the web. This exists as a single place to eyeball the data, and as a possible
    Power Query source.
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    out = os.path.expanduser(args.out)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary = []

    for slug in (args.datasets or pivot.list_specs()):
        spec = pivot.load_spec(slug)
        fn = sources.SOURCES[spec['source']]
        rows, meta = fn(month=args.month)
        built = pivot.build(rows, spec)
        title = spec.get('sheet_name', slug.replace('-', ' ').upper())[:31]
        ws = wb.create_sheet(title)

        for r in pivot.preamble_rows(spec, meta):
            ws.append(r)
        ws.append(built['group_row'])
        ws.append(built['header_row'])
        hdr_rows = (ws.max_row - 1, ws.max_row)
        for line in built['grid']:
            ws.append(line)

        bold = Font(bold=True)
        for rr in hdr_rows:
            for c in ws[rr]:
                c.font = bold
        ws.freeze_panes = ws.cell(row=hdr_rows[1] + 1, column=len(spec.get('fixed', [])) + 1)
        ws.column_dimensions['A'].width = 26
        # first column of each warehouse block gets a little room for the label
        for i in range(len(spec.get('fixed', [])) + 1, len(built['header_row']) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 13

        summary.append((title, len(built['grid']), len(built['groups']), meta.get('period') or '-'))
        print(f"  {title:<16} {len(built['grid']):>6} rows x {len(built['groups'])} warehouses"
              f"  {meta.get('period') or ''}")

    idx = wb.create_sheet('README', 0)
    idx.append(['Rapid Labels — Excel Sync master'])
    idx.append([f'Built {datetime.now().strftime("%d-%b-%Y %H:%M")} from cin7_mirror via Supabase.'])
    idx.append([])
    idx.append(['Tab', 'Rows', 'Warehouses', 'Period'])
    for row in summary:
        idx.append(list(row))
    idx.append([])
    idx.append(['Source of truth is excel_sync.dataset_rows in Supabase, not this file.'])
    idx.append(['The branch workbooks are written directly and do NOT read from here.'])
    for c in idx[1] + idx[4]:
        c.font = Font(bold=True)
    idx.column_dimensions['A'].width = 62
    for col in 'BCD':
        idx.column_dimensions[col].width = 14

    os.makedirs(os.path.dirname(out) or '.', exist_ok=True)
    wb.save(out)
    print(f"\n  wrote {out}")
    return 0


def main(argv=None):
    p = argparse.ArgumentParser(prog='excel-sync')
    sub = p.add_subparsers(dest='cmd', required=True)
    sub.add_parser('list', help='show datasets and the workbook bindings').set_defaults(fn=cmd_list)
    sub.add_parser('register', help='mirror bindings into ops.sync_registry').set_defaults(fn=cmd_register)

    b = sub.add_parser('build', help='build a dataset from the mirror')
    b.add_argument('slug')
    b.add_argument('--month', help='YYYY-MM (defaults to the current Sydney month)')
    b.add_argument('--out', help='directory to write csv/xlsx into')
    b.add_argument('--verify', help='Cin7 export to diff against')
    b.add_argument('--publish', action='store_true',
                   help='store the dataset in Supabase for the workbooks to read')
    b.add_argument('--force', action='store_true',
                   help='publish even if a gate failed (records the run as blocked)')
    b.add_argument('--show', type=int, default=8, help='how many differing cells to print')
    b.add_argument('--tolerance', type=float, default=1.0, help='max %% off the export')
    b.set_defaults(fn=cmd_build)

    r = sub.add_parser('render', help='build a workbook tab block and diff it against the tab')
    r.add_argument('slug')
    r.add_argument('--workbook', help='path to the .xlsx to diff against')
    r.add_argument('--out', help='directory to write the block as csv')
    r.add_argument('--show', type=int, default=8)
    r.set_defaults(fn=cmd_render)

    sub.add_parser('graph-login',
                   help='sign in once so the nightly run can refresh its own token'
                   ).set_defaults(fn=cmd_graph_login)

    d = sub.add_parser('deliver', help='write a binding into its workbook tab')
    d.add_argument('slug', nargs='?', help='binding slug (omit for all bindings)')
    # Writing is opt-in, and it is the only irreversible thing this repo does.
    d.add_argument('--write', action='store_true',
                   help='actually write. Without it this is a rehearsal that proves '
                        'every gate and touches nothing.')
    d.add_argument('--file', help='write to THIS workbook instead of the binding\'s — '
                                  'how a single binding is aimed at one copy')
    d.add_argument('--root', help='look for every workbook under THIS folder instead of the '
                                  'synced library — how all 21 bindings are aimed at a folder '
                                  'of test copies at once. Falls back to <root>/<file> when '
                                  'the library subfolders are not reproduced.')
    d.add_argument('--force', action='store_true',
                   help='ignore the disabled flag and the staleness gate. Does NOT override '
                        'the header gate or the formula guard — those two protect against '
                        'silent corruption and have no legitimate override.')
    d.add_argument('--mode', choices=['auto', 'app', 'delegated'], default=None,
                   help='graph only: force an auth door (default: app-only if possible)')
    d.add_argument('--transport', choices=['local', 'graph'], default=None,
                   help="'local' drives Excel over the OneDrive-synced copy and needs no "
                        "tenant permission; 'graph' writes to SharePoint over HTTP and is "
                        'blocked until an admin consents. Default: EXCEL_SYNC_TRANSPORT.')
    d.set_defaults(fn=cmd_deliver)
    a = sub.add_parser('apply', help='write the blocks into a workbook on disk')
    a.add_argument('workbook', help='path to the .xlsx to write')
    a.add_argument('--bindings', nargs='+', required=True, help='binding slugs to apply')
    a.add_argument('--backup-dir', default='out/backups')
    a.add_argument('--dry-run', action='store_true')
    a.add_argument('--i-know-this-is-live', action='store_true',
                   help='required to write inside a cloud-synced folder')
    a.set_defaults(fn=cmd_apply)

    m = sub.add_parser('master', help='one workbook with every dataset as a tab')
    m.add_argument('--out', default='out/RapidLED-Master.xlsx')
    m.add_argument('--datasets', nargs='+')
    m.add_argument('--month')
    m.set_defaults(fn=cmd_master)

    args = p.parse_args(argv)
    return args.fn(args)


if __name__ == '__main__':
    sys.exit(main())
