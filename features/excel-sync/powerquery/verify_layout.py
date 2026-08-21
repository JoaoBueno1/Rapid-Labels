"""Largura, formato numerico e carimbo nas 19 abas, contra o backup."""
import openpyxl, os, json, glob, warnings, collections
warnings.filterwarnings('ignore')
from openpyxl.utils import get_column_letter as CL
DIR = os.path.expanduser('~/OneDrive - RapidLED/Desktop/Tests files')
BK  = os.path.abspath(sorted(glob.glob('BACKUP_*'))[-1])
B   = [b for b in json.load(open('/tmp/bindings.json')) if 'Hobart' not in b['file']]

def widths(ws, n=20):
    out = {}
    for c in range(1, n+1):
        d = ws.column_dimensions.get(CL(c))
        out[CL(c)] = round(d.width, 2) if (d and d.width) else None
    return out

nw = nf = 0
for fn in sorted({b['file'] for b in B}):
    wo = openpyxl.load_workbook(os.path.join(BK, fn))
    wn = openpyxl.load_workbook(os.path.join(DIR, fn))
    wv = openpyxl.load_workbook(os.path.join(DIR, fn), data_only=True)
    print(f'\n{fn}')
    for b in [x for x in B if x['file'] == fn]:
        o, n_ = wo[b['sheet']], wn[b['sheet']]
        a = n_[b['anchor']]; c0, r0 = a.column, a.row
        ncol = len(b['cols'])
        dw = {k: (v, widths(n_)[k]) for k, v in widths(o).items() if v != widths(n_)[k]}
        bad = [f'{n_.cell(r,c).coordinate}={n_.cell(r,c).number_format}'
               for r in range(r0, min(r0+4000, n_.max_row+1))
               for c in range(c0, c0+ncol)
               if n_.cell(r,c).number_format not in ('General','@')]
        nw += len(dw); nf += len(bad)
        st = str(wv[b['sheet']][b['cell']].value or '')
        print(f'   {b["sheet"]:<11} largura {"OK" if not dw else dw}  '
              f'formatos estranhos {len(bad)}')
        print(f'        {st[:82]}')
    wo.close(); wn.close(); wv.close()
print(f'\n  larguras alteradas: {nw}   formatos estranhos: {nf}')
