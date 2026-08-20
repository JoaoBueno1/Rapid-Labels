"""Quem atualizou, quando, e esta tudo inteiro? Le os arquivos, nao os toca.

Nao precisa de nada novo no banco nem nas planilhas: cada arquivo ja carrega a
hora do proprio refresh na aba _Sync, e o zip diz quem gravou por ultimo.
"""
import openpyxl, os, re, sys, zipfile, datetime, warnings, json
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

LOCAIS = [
  ('filiais', 'C:/Users/JoaoMarcos/RapidLED/WorkDocs - Rapid LED - Data/'
              'Inventory Management/Inventory Stock Orders'),
  ('gateway', 'C:/Users/JoaoMarcos/RapidLED/Inventory Management - Documents/Gateway'),
]
MESES = {m: i for i, m in enumerate(
    ['January','February','March','April','May','June','July','August',
     'September','October','November','December'], 1)}

def quando(txt):
    """'Thursday, 20 August 2026 at 2:53 pm (Brisbane time)' -> datetime"""
    m = re.search(r'(\d{1,2}) (\w+) (\d{4}) at (\d{1,2}):(\d{2}) (am|pm)', str(txt or ''))
    if not m: return None
    d, mes, a, h, mi, ap = m.groups()
    h = int(h) % 12 + (12 if ap == 'pm' else 0)
    return datetime.datetime(int(a), MESES[mes], int(d), h, int(mi))

def olhar(p):
    r = {'arquivo': os.path.basename(p), 'tamanho_mb': os.path.getsize(p)/1e6}
    r['gravado_em'] = datetime.datetime.fromtimestamp(os.path.getmtime(p))
    try:
        with zipfile.ZipFile(p) as z:
            app = re.search(r'<Application>([^<]*)</Application>',
                            z.read('docProps/app.xml').decode('utf8','replace'))
            r['gravado_por'] = app.group(1) if app else '?'
            r['consultas'] = len([x for x in z.namelist() if 'queryTable' in x])
        # read_only + iteracao por linha: varrer max_row em arquivo de 4 MB
        # levava minutos e o monitor tem de rodar em segundos.
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        r['abas'] = len(wb.sheetnames)
        r['tem_sync'] = '_Sync' in wb.sheetnames
        r['refresh'] = r['dado_de'] = None
        if r['tem_sync']:
            linhas = list(wb['_Sync'].iter_rows(min_row=1, max_row=2, values_only=True))
            if len(linhas) > 1:
                v = linhas[1]
                r['refresh'] = quando(v[0] if len(v) > 0 else None)
                r['dado_de'] = quando(v[2] if len(v) > 2 else None)
        vazias = []
        for nome in wb.sheetnames:
            if nome.startswith(('SOH', 'Sales MTD', 'Stock Data', 'Restock')):
                ws = wb[nome]
                achou = False
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=12, values_only=True)):
                    if any(c not in (None, '') for c in (row or ())[:6]): achou = True; break
                if not achou: vazias.append(nome)
        r['abas_vazias'] = vazias
        wb.close()
    except PermissionError:
        r['erro'] = 'ABERTO NO EXCEL'
    except Exception as e:
        r['erro'] = type(e).__name__
    return r

if __name__ == '__main__':
    agora = datetime.datetime.now()
    print(f'{"arquivo":<28}{"ultimo refresh":<17}{"idade":>7}  {"gravado por":<22}{"cons":>5} {"aviso"}')
    for _, d in LOCAIS:
        if not os.path.isdir(d): continue
        for fn in sorted(os.listdir(d)):
            if not fn.endswith('.xlsx') or fn.startswith('~$'): continue
            r = olhar(os.path.join(d, fn))
            if r.get('erro'):
                print(f'  {r["arquivo"]:<26}{r["erro"]}'); continue
            ref = r['refresh']
            idade = f'{(agora-ref).total_seconds()/3600:5.1f}h' if ref else '   ---'
            avisos = []
            if not r['tem_sync']: avisos.append('sem _Sync')
            if ref and (agora-ref).days >= 1: avisos.append('nao atualiza ha %dd' % (agora-ref).days)
            if r['abas_vazias']: avisos.append('VAZIA: ' + ','.join(r['abas_vazias']))
            print(f'  {r["arquivo"]:<26}{ref.strftime("%d/%m %H:%M") if ref else "-":<17}'
                  f'{idade:>7}  {r["gravado_por"]:<22}{r["consultas"]:>5} '
                  f'{"; ".join(avisos) or "ok"}')
