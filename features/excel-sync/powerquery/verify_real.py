"""Verifica um arquivo REAL contra a fonte e contra o backup pre-migracao."""
import os, sys, json, glob, warnings, re, urllib.request, urllib.parse, openpyxl
warnings.filterwarnings('ignore')
from openpyxl.utils import get_column_letter as CL
REAL = 'C:/Users/JoaoMarcos/RapidLED/WorkDocs - Rapid LED - Data/Inventory Management/Inventory Stock Orders'
BK = os.path.abspath(sorted(glob.glob('BACKUP_REAL_*'))[-1])
env = {}
for line in open(os.path.expanduser('~/Rapid-Labels/.env'), encoding='utf8'):
    m = re.match(r'\s*([A-Z0-9_]+)\s*=\s*(.*)', line)
    if m: env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
URL, KEY = env['SUPABASE_URL'], env['SUPABASE_ANON_KEY']

def ref(b):
    f = 'in.(' + ','.join('"%s"' % l for l in b['locs']) + ')'
    rows, off = [], 0
    while True:
        q = urllib.parse.urlencode({'p_dataset':b['dataset'],'location':f,
                                    'limit':'1000','offset':str(off)})
        r = urllib.request.Request(f'{URL}/rest/v1/rpc/excel_dataset_rows?{q}')
        r.add_header('apikey',KEY); r.add_header('Authorization','Bearer '+KEY)
        d = json.load(urllib.request.urlopen(r, timeout=180))
        if not d: break
        rows += d; off += 1000
    acc = {}
    for x in rows:
        a = acc.setdefault(x['sku'], {})
        for k,v in (x['metrics'] or {}).items():
            a[k] = a.get(k,0.0) + (0.0 if v in (None,'') else float(v))
    out = []
    for s in sorted(acc, key=lambda z:(z.lower(), z)):
        row=[s]
        for c in b['cols'][1:]:
            row.append('' if c[2] else (round(acc[s].get(c[0],0.0), c[3])
                       if c[3] is not None else acc[s].get(c[0],0.0)))
        out.append(row)
    return out

def same(a,b):
    if a in (None,'') and b in (None,''): return True
    try: return abs(float(a or 0)-float(b or 0)) < 1e-6
    except (TypeError,ValueError): return str(a or '')==str(b or '')

def widths(ws, n=16):
    return {CL(c): (round(ws.column_dimensions[CL(c)].width,2)
            if CL(c) in ws.column_dimensions and ws.column_dimensions[CL(c)].width else None)
            for c in range(1,n+1)}

def check(fn):
    B = [b for b in json.load(open('/tmp/bindings.json')) if b['file']==fn]
    wv = openpyxl.load_workbook(os.path.join(REAL,fn), data_only=True)
    wn = openpyxl.load_workbook(os.path.join(REAL,fn))
    wo = openpyxl.load_workbook(os.path.join(BK,fn))
    print(f'\n=== {fn} ===')
    print(f'  abas {len(wn.sheetnames)}   _Sync: {wn["_Sync"].sheet_state}')
    allok = True
    for b in B:
        ws = wv[b['sheet']]; a = ws[b['anchor']]; c0,r0=a.column,a.row; n=len(b['cols'])
        e = ref(b); got=[]; i=0
        while ws.cell(r0+i,c0).value not in (None,''):
            got.append([ws.cell(r0+i,c0+j).value for j in range(n)]); i+=1
        bad=[1 for g,x in zip(got,e) for j in range(n) if not same(g[j],x[j])]
        ok = len(got)==len(e) and not bad; allok &= ok
        w0, w1 = widths(wo[b['sheet']]), widths(wn[b['sheet']])
        dw = {k:(v,w1[k]) for k,v in w0.items() if v!=w1[k]}
        nf = sum(1 for r in range(r0, min(r0+4000, wn[b['sheet']].max_row+1))
                 for c in range(c0, c0+n)
                 if wn[b['sheet']].cell(r,c).number_format not in ('General','@'))
        print(f'  {b["sheet"]:<11}{len(got):>5} vs {len(e):>5} fonte  '
              f'{"IDENTICO" if ok else "DIVERGE("+str(len(bad))+")":<12} '
              f'largura {"OK" if not dw else dw}  formatos {nf}')
        print(f'       {str(ws[b["cell"]].value)[:76]}')
    wv.close(); wn.close(); wo.close()
    return allok

if __name__ == '__main__':
    r = [check(f) for f in sys.argv[1:]]
    print(f'\n  {sum(r)}/{len(r)} arquivos conferem')
