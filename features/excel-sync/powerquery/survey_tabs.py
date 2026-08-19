import openpyxl, json, re, os, sys
from openpyxl.utils import get_column_letter as CL, column_index_from_string as CI

DIR = os.path.expanduser('~/OneDrive - RapidLED/Desktop/Tests files')
B = json.load(open('/tmp/bindings.json'))
B = [b for b in B if 'Hobart' not in b['file']]

def split(a):
    m = re.match(r'^([A-Z]+)(\d+)$', a.upper()); return CI(m.group(1)), int(m.group(2))

out = []
for fn in sorted({b['file'] for b in B}):
    wb = openpyxl.load_workbook(os.path.join(DIR, fn), data_only=False)
    print(f'\n{"="*78}\n{fn}   ({len(wb.sheetnames)} abas)')
    for b in [x for x in B if x['file'] == fn]:
        ws = wb[b['sheet']]
        c0, r0 = split(b['anchor']); ncol = len(b['cols'])
        hdr_r = r0 - 1
        hdrs = [ws.cell(hdr_r, c0 + i).value for i in range(ncol)]
        last = r0 - 1
        for r in range(r0, ws.max_row + 1):
            if ws.cell(r, c0).value not in (None, ''): last = r
        nrows = last - r0 + 1

        # formulas dentro do retangulo de dados?
        fx = [f'{CL(c0+i)}{r}' for r in range(r0, min(last, r0+4000)+1)
              for i in range(ncol)
              if isinstance(ws.cell(r, c0+i).value, str) and str(ws.cell(r, c0+i).value).startswith('=')]
        # o que existe A DIREITA das colunas de dados, na faixa de linhas da tabela
        right = []
        for r in list(range(hdr_r, min(hdr_r+6, ws.max_row+1))) + [last]:
            for c in range(c0 + ncol, min(c0 + ncol + 8, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v not in (None, ''): right.append(f'{CL(c)}{r}={str(v)[:34]!r}')
        merged = [str(m) for m in ws.merged_cells.ranges
                  if m.min_row >= hdr_r - 1 and m.min_col >= c0 - 1 and m.min_col <= c0 + ncol + 4]
        cf = list(getattr(ws.conditional_formatting, '_cf_rules', {}) or {})
        rec = dict(file=fn, sheet=b['sheet'], anchor=b['anchor'], hdr_row=hdr_r,
                   col0=c0, ncol=ncol, headers=hdrs, expected=b['cols'],
                   first_row=r0, last_row=last, nrows=nrows, formulas=fx[:6],
                   n_formulas=len(fx), right=right[:10], merged=merged[:6],
                   n_cf=len(cf), freeze=ws.freeze_panes, max_col=ws.max_column,
                   row1=[f'{CL(c)}1={str(ws.cell(1,c).value)[:30]!r}'
                         for c in range(1, min(12, ws.max_column+1))
                         if ws.cell(1,c).value not in (None,'')])
        out.append(rec)
        ok = [h == e[1] for h, e in zip(hdrs, b['cols'])]
        print(f'  {b["sheet"]:<11} {b["anchor"]:<4} hdr=L{hdr_r} {nrows:>5} linhas '
              f'({CL(c0)}{r0}:{CL(c0+ncol-1)}{last})  cabecalho {"OK" if all(ok) else "DIVERGE"}')
        if not all(ok):
            print(f'      encontrado: {hdrs}')
            print(f'      esperado  : {[e[1] for e in b["cols"]]}')
        if fx: print(f'      !! {len(fx)} FORMULAS no retangulo: {fx[:3]}')
        if merged: print(f'      merged perto: {merged[:3]}')
        if right: print(f'      a direita: {right[:5]}')
    wb.close()

json.dump(out, open('/tmp/tabs.json','w'), indent=1, default=str)
print(f'\n{len(out)} abas -> /tmp/tabs.json')
