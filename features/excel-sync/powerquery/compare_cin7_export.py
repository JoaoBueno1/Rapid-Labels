"""Compara a aba da planilha contra o export do Cin7, SKU a SKU.

Nao basta bater a contagem: o que importa e quem esta de um lado e nao do outro.
"""
import openpyxl, os, glob, warnings, json
warnings.filterwarnings('ignore')
DL  = os.path.expanduser('~/Downloads')
DIR = os.path.expanduser('~/OneDrive - RapidLED/Desktop/Tests files')

def num(v):
    if v in (None, ''): return 0.0
    try: return float(v)
    except (TypeError, ValueError): return 0.0

def read_stock(path):
    """L1=location, L2=cabecalho, L3+=dados. Soma quando ha varias locations."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True)); wb.close()
    locs = [str(c).strip() if c else '' for c in rows[0]]
    hdr  = [str(c).strip() if c else '' for c in rows[1]]
    want = {'Quantity on hand':'on_hand', 'Allocated':'allocated',
            'On order':'on_order', 'Available':'available'}
    cols = [(i, want[h]) for i, h in enumerate(hdr) if h in want]
    out = {}
    for r in rows[2:]:
        if not r or not r[0]: continue
        sku = str(r[0]).strip()
        a = out.setdefault(sku, {v: 0.0 for v in want.values()})
        for i, f in cols:
            if i < len(r): a[f] += num(r[i])
    return out, sorted({l for l in locs if l})

def read_sales(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True)); wb.close()
    hi = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == 'SKU')
    hdr = [str(c).strip() if c else '' for c in rows[hi]]
    want = {'Quantity':'quantity', 'Total':'total'}
    cols = [(i, want[h]) for i, h in enumerate(hdr) if h in want]
    out = {}
    for r in rows[hi+1:]:
        if not r or not r[0]: continue
        sku = str(r[0]).strip()
        a = out.setdefault(sku, {v: 0.0 for v in want.values()})
        for i, f in cols:
            if i < len(r): a[f] += num(r[i])
    return out, [str(rows[hi-1][1]).strip()] if hi else []

def read_tab(fn, sheet, anchor, fields):
    wb = openpyxl.load_workbook(os.path.join(DIR, fn), data_only=True)
    ws = wb[sheet]; a = ws[anchor]; c0, r0 = a.column, a.row
    out, i = {}, 0
    while True:
        sku = ws.cell(r0+i, c0).value
        if sku in (None, ''): break
        out[str(sku).strip()] = {f: num(ws.cell(r0+i, c0+1+j).value)
                                 for j, f in enumerate(fields)}
        i += 1
    wb.close(); return out

def compare(label, cin7, tab, fields):
    only_c = sorted(set(cin7) - set(tab))
    only_t = sorted(set(tab) - set(cin7))
    diff = []
    for s in sorted(set(cin7) & set(tab)):
        for f in fields:
            if abs(cin7[s][f] - tab[s][f]) > 0.005:
                diff.append((s, f, cin7[s][f], tab[s][f]))
    print(f'\n  {label}')
    print(f'     Cin7 {len(cin7):>5} SKUs   planilha {len(tab):>5} SKUs   '
          f'em comum {len(set(cin7)&set(tab)):>5}')
    print(f'     so no Cin7: {len(only_c):>4}   so na planilha: {len(only_t):>4}   '
          f'valores diferentes: {len(diff)}')
    for s in only_c[:5]: print(f'        so Cin7     {s:<34} {cin7[s]}')
    for s in only_t[:5]: print(f'        so planilha {s:<34} {tab[s]}')
    for s, f, a, b in diff[:6]:  print(f'        difere {s:<26} {f:<10} Cin7={a:g} planilha={b:g}')
    return only_c, only_t, diff

if __name__ == '__main__':
    SF = ['on_hand','allocated','on_order','available']
    g = lambda pat: sorted(glob.glob(os.path.join(DL, pat)))[-1]
    print('=== ESTOQUE ===')
    c, l = read_stock(g('Inventory Products Stock Level Report - 2026-08-20T070947*.xlsx'))
    compare(f'Coffs Harbour / SOH Dear   (export: {l})', c,
            read_tab('Coffs Harbour Aug 26.xlsx','SOH Dear','B3',SF), SF)
    c, l = read_stock(g('Inventory Products Stock Level Report - 2026-08-20T070903*.xlsx'))
    compare(f'Sunshine Coast / SOH SC    (export: {l})', c,
            read_tab('Sunshine Coast Aug 26.xlsx','SOH SC','B3',SF), SF)
    c, l = read_stock(g('Inventory Products Stock Level Report - 2026-08-20T071015*.xlsx'))
    compare(f'Main+Gateway / SOH Main    (export: {l})', c,
            read_tab('Coffs Harbour Aug 26.xlsx','SOH Main','B3',SF), SF)
    print('\n=== VENDAS DO MES ===')
    QF = ['quantity','total']
    c, l = read_sales(g('Sale Order Details - 2026-08-20T071101*.xlsx'))
    tab = read_tab('Coffs Harbour Aug 26.xlsx','Sales MTD','A7',['quantity','discount','total'])
    compare(f'Coffs Harbour / Sales MTD  (export: {l})', c,
            {k: {'quantity': v['quantity'], 'total': v['total']} for k, v in tab.items()}, QF)
