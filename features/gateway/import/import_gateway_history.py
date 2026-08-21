"""Migrate the Gateway lot ledger out of 'Gateway Driver Aug 26.xlsx'.

The source is the workbook's 'MAIN Stock Movement' sheet, which — despite
being a spreadsheet — is already a lot ledger. One row is one pallet of one
SKU on one shelf, and its columns are laid out as events:

    C Shelf Alloc  D Pallet #  E SKU  F CURRENTY QTY
    G Date   H First Qty  I Transfer#      <- the receipt that created the lot
    J Date   K Unit Qty   L Transfer       <- withdrawal 1
    M Date   N Unit Qty   O Transfer       <- withdrawal 2
    P Date   Q Unit Qty   R Transfer       <- withdrawal 3

That is exactly a receipt plus its drawdowns, which is why this migration is a
translation rather than a reconstruction. The spreadsheet's limit is that there
are only three withdrawal slots: a fourth overwrites the first, so the sheet
cannot be replayed past three moves per pallet. The new ledger has no such cap.

Which number is the balance
---------------------------
Measured against Cin7's own Gateway totals, over the 387 SKUs present in both:

    CURRENTY QTY (hand-keyed)      328 of 387 exact (84.8%)   gap  4,010 units
    First Qty - withdrawals        186 of 387 exact (48.1%)   gap 47,743 units

so the hand-typed running total anchors the balance, not the arithmetic. That
is counter-intuitive — the arithmetic looks more trustworthy — but 'First Qty'
is blank on 292 rows that still hold stock, and the sheet has only three
withdrawal slots, so replaying it loses every movement past the third.

The withdrawals still carry real dates and real TR references, so they are
replayed against a receipt of (CURRENTY QTY + everything replayed). The
remaining balance therefore lands on CURRENTY QTY by construction, and
qty_received is explicitly derived: what must have arrived for the recorded
withdrawals to leave today's balance. Where the row's own 'First Qty'
contradicts that, BOTH numbers go into gateway_import_issues.

What this script will NOT do
----------------------------
It will not invent a date, a quantity or a transfer reference. A cell it cannot
read becomes an issue for a human, never a plausible-looking value.

Usage
-----
    python features/gateway/import/import_gateway_history.py --dry-run
    python features/gateway/import/import_gateway_history.py --apply
    python features/gateway/import/import_gateway_history.py --apply --replace
    python features/gateway/import/import_gateway_history.py --rollback <batch_id>

Needs SUPABASE_URL and SUPABASE_SERVICE_KEY (read from ../../.env).
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from typing import Any

import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
DEFAULT_WORKBOOK = os.path.expanduser(
    r"~\RapidLED\Inventory Management - Documents\Gateway\Gateway Driver Aug 26.xlsx")
SHEET = "MAIN Stock Movement"
HEADER_ROW = 3          # real header; data starts at row 4
KIND = "lot_ledger"

# Column indices (0-based) on 'MAIN Stock Movement'.
C_ALERT, C_5DC, C_SHELF, C_PALLET, C_SKU, C_CURRENT = 0, 1, 2, 3, 4, 5
C_RECV_DATE, C_FIRST_QTY, C_RECV_TR = 6, 7, 8
OUT_SLOTS = ((9, 10, 11), (12, 13, 14), (15, 16, 17))   # (date, qty, transfer)
C_NOTES, C_LOC_EXPANDED = 20, 21

# Anything outside this window is a typo or a stray Excel serial, not a date.
# The sheet really does contain 1900-01-16, 1936-11-28 and 1941-09-21.
DATE_MIN = dt.date(2015, 1, 1)


def _env() -> tuple[str, str]:
    env: dict[str, str] = {}
    path = os.path.join(REPO, ".env")
    if os.path.exists(path):
        for line in open(path, encoding="utf8"):
            m = re.match(r"\s*([A-Z0-9_]+)\s*=\s*(.*)", line)
            if m:
                env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
    url = os.environ.get("SUPABASE_URL") or env.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_KEY") or env.get("SUPABASE_SERVICE_KEY", "")
    if not url or not key:
        sys.exit("SUPABASE_URL / SUPABASE_SERVICE_KEY not found in environment or .env")
    return url.rstrip("/"), key


class Supa:
    def __init__(self, url: str, key: str):
        self.url, self.key = url, key

    def _req(self, method: str, path: str, body: Any = None, headers: dict | None = None):
        data = json.dumps(body).encode() if body is not None else None
        req = urllib.request.Request(f"{self.url}{path}", data=data, method=method)
        req.add_header("apikey", self.key)
        req.add_header("Authorization", f"Bearer {self.key}")
        req.add_header("Content-Type", "application/json")
        for k, v in (headers or {}).items():
            req.add_header(k, v)
        try:
            with urllib.request.urlopen(req, timeout=180) as r:
                raw = r.read().decode()
                return json.loads(raw) if raw.strip() else None
        except urllib.error.HTTPError as e:
            sys.exit(f"{method} {path} -> HTTP {e.code}: {e.read().decode()[:1200]}")

    def select(self, table: str, query: str = ""):
        return self._req("GET", f"/rest/v1/{table}?{query}")

    def insert(self, table: str, rows: Any):
        return self._req("POST", f"/rest/v1/{table}", rows, {"Prefer": "return=representation"})

    def patch(self, table: str, query: str, body: Any):
        return self._req("PATCH", f"/rest/v1/{table}?{query}", body, {"Prefer": "return=representation"})

    def rpc(self, fn: str, args: dict):
        return self._req("POST", f"/rest/v1/rpc/{fn}", args)


# ── cell readers ────────────────────────────────────────────────────────────
def _blank(v: Any) -> bool:
    """Excel hands back U+00A0 for cells that merely LOOK empty — the whole
    'Date Expected' column is nothing but non-breaking spaces. Treating those
    as content turns 200-odd empty cells into fake parse failures."""
    if v is None:
        return True
    if isinstance(v, str):
        return v.replace("\xa0", " ").strip() == ""
    return False


def read_date(v: Any) -> tuple[dt.date | None, str | None]:
    """Return (date, problem). A problem never becomes a date."""
    if _blank(v):
        return None, None
    if isinstance(v, dt.datetime):
        d = v.date()
    elif isinstance(v, dt.date):
        d = v
    else:
        # '17-07', '23-11', 'extrusion', '15-ju' all live in date-formatted cells
        return None, f"unparseable date {v!r}"
    if d < DATE_MIN:
        return None, f"date {d.isoformat()} is before {DATE_MIN.isoformat()}"
    if d > dt.date.today() + dt.timedelta(days=30):
        return None, f"date {d.isoformat()} is in the future"
    return d, None


TR_LIKE = re.compile(r"^\s*tr-?\d+\s*-?\s*$", re.I)


def read_qty(v: Any) -> tuple[float | None, str | None]:
    if _blank(v):
        return None, None
    if isinstance(v, bool):
        return None, f"boolean in a quantity cell: {v!r}"
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).replace("\xa0", " ").strip()
    if TR_LIKE.match(s):
        # A transfer reference typed one column to the left of where it
        # belongs. The reference is recoverable; the quantity is simply gone.
        return None, f"COLUMN SHIFT: transfer reference {s!r} is in the quantity column"
    # '51 CTN P2 TOTAL 102' — a real value in the First Qty column. Refuse to
    # guess which of the two numbers is the quantity.
    return None, f"non-numeric quantity {s[:60]!r}"


XL_ERRORS = {"#REF!", "#N/A", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!"}


def clean(v: Any) -> str | None:
    """Excel error values arrive as their literal text. 36 cells of the
    'Location Expanded' column evaluate to #REF!, and a shelf called '#REF!'
    is worse than no shelf at all."""
    if _blank(v):
        return None
    s = str(v).replace("\xa0", " ").strip()
    if s.upper() in XL_ERRORS:
        return None
    return s or None


def read_tr(v: Any) -> str | None:
    s = clean(v)
    if not s:
        return None
    s = s.upper().replace(" ", "").rstrip("-")
    # 'tr-31534' and 'TR-30573' sit in adjacent cells
    m = re.fullmatch(r"TR-?(\d+)", s)
    return f"TR-{m.group(1)}" if m else s


def read_5dc(v: Any) -> str | None:
    """SAP stores Item No. as text for pasted rows and as a number for typed
    ones, so the same code arrives as '30313' or 30313. '0' means the VLOOKUP
    found the SKU but its Item No. cell was blank."""
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return None if s in ("0", "#N/A", "") else s


def load_sku_index(sb: "Supa") -> dict[str, str]:
    """lower(sku) -> the spelling Cin7 actually uses.

    People type '12V-IP20-030W'; Cin7 holds '12v-IP20-030w'. Upper-casing on
    import would leave 64 SKUs unable to join cin7_mirror.products, and the
    reconciliation screen would then invent a discrepancy for every one of
    them. So the workbook's free-text SKU is resolved to Cin7's own spelling
    before anything is written.
    """
    index: dict[str, str] = {}
    ambiguous: set[str] = set()
    offset = 0
    while True:
        rows = sb._req(
            "GET",
            f"/rest/v1/products?select=sku&order=sku.asc&limit=1000&offset={offset}",
            headers={"Accept-Profile": "cin7_mirror"}) or []
        if not rows:
            break
        for r in rows:
            k = (r["sku"] or "").lower()
            if k in index and index[k] != r["sku"]:
                ambiguous.add(k)
            index.setdefault(k, r["sku"])
        offset += 1000
    for k in ambiguous:            # two products differing only by case: refuse to pick
        index.pop(k, None)
    return index


# ── parse ───────────────────────────────────────────────────────────────────
def parse_workbook(path: str, sku_index: dict[str, str] | None = None) -> tuple[list[dict], list[dict], dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"sheet {SHEET!r} not found in {path}")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()

    parsed: list[dict] = []
    issues: list[dict] = []
    stats = Counter()
    shelf_carry: str | None = None

    def issue(sev, code, row_no, sku, msg, raw=None):
        issues.append({"severity": sev, "code": code, "sheet": SHEET,
                       "row_ref": f"{SHEET}!r{row_no}", "sku": sku,
                       "message": msg, "raw": raw})
        stats[f"issue_{sev}"] += 1

    for idx in range(HEADER_ROW, len(rows)):
        row = rows[idx]
        row_no = idx + 1
        if not any(x not in (None, "") for x in row[:22]):
            continue

        sku = clean(row[C_SKU])
        # Column C is filled only on the first row of a merged shelf block;
        # continuation rows inherit it. Column V is the sheet's own fill-down
        # of exactly that, so prefer it and fall back to carrying C ourselves.
        shelf_here = clean(row[C_SHELF])
        if shelf_here:
            shelf_carry = shelf_here
        shelf = clean(row[C_LOC_EXPANDED]) or shelf_carry

        if not sku:
            continue
        stats["rows_with_sku"] += 1
        raw_sku = sku
        if sku_index is not None:
            canonical = sku_index.get(sku.lower())
            if canonical:
                sku = canonical
                if canonical != raw_sku:
                    stats["sku_recased"] += 1
            else:
                stats["sku_unknown"] += 1
                issue("warning", "sku_not_in_cin7", row_no, raw_sku,
                      f"{raw_sku!r} does not exist in cin7_mirror.products — lot imported "
                      f"with the SKU exactly as typed, and it will show as local-only in "
                      f"reconciliation until the product is created or the SKU corrected")

        recv_date, date_problem = read_date(row[C_RECV_DATE])
        first_qty, qty_problem = read_qty(row[C_FIRST_QTY])
        current_qty, _ = read_qty(row[C_CURRENT])
        row_ref = f"{SHEET}!r{row_no}"

        if date_problem:
            issue("warning", "bad_receipt_date", row_no, sku,
                  f"{date_problem} — lot imported without a receipt date",
                  {"cell": f"G{row_no}", "value": str(row[C_RECV_DATE])[:80]})
        if qty_problem:
            issue("warning", "bad_first_qty", row_no, sku,
                  f"{qty_problem} — falling back to CURRENTY QTY as an opening balance",
                  {"cell": f"H{row_no}", "value": str(row[C_FIRST_QTY])[:80]})

        # withdrawals
        outs, out_total, out_dates_missing = [], 0.0, 0
        for slot, (dc, qc, tc) in enumerate(OUT_SLOTS, start=1):
            q, qp = read_qty(row[qc])
            if q is None or q <= 0:
                if qp:
                    shifted = qp.startswith("COLUMN SHIFT")
                    issue("warning",
                          "transfer_ref_in_qty_column" if shifted else "bad_withdrawal_qty",
                          row_no, sku, f"withdrawal {slot}: {qp} — no movement recorded",
                          {"cell": f"{chr(65+qc)}{row_no}", "value": str(row[qc])[:80],
                           "recovered_reference": read_tr(row[qc]) if shifted else None})
                continue
            d, dp = read_date(row[dc])
            if dp:
                out_dates_missing += 1
                issue("warning", "bad_withdrawal_date", row_no, sku,
                      f"withdrawal {slot}: {dp} — dated to the receipt instead",
                      {"cell": f"{chr(65+dc)}{row_no}", "value": str(row[dc])[:80]})
            if d and recv_date and d < recv_date:
                issue("warning", "withdrawal_before_receipt", row_no, sku,
                      f"withdrawal {slot} is dated {d.isoformat()} but the receipt is "
                      f"{recv_date.isoformat()} — kept as recorded, the order is wrong",
                      {"withdrawal_date": d.isoformat(), "receipt_date": recv_date.isoformat()})
            outs.append({
                "occurred_at": (dt.datetime.combine(d, dt.time(0, 0)).isoformat() if d else None),
                "qty": q, "reference": read_tr(row[tc])})
            out_total += q

        # ── How much stock does this row represent? ──────────────────────
        #
        # Measured against Cin7's own Gateway totals across the 387 SKUs that
        # appear in both, per SKU exact matches / total absolute unit gap:
        #
        #     CURRENTY QTY (hand-keyed)      328 of 387 (84.8%)   gap  4,010
        #     First Qty - withdrawals        186 of 387 (48.1%)   gap 47,743
        #
        # The hand-typed running total wins by an order of magnitude, because
        # 'First Qty' is blank on 292 rows that still hold stock and the sheet
        # only has three withdrawal slots, so the arithmetic silently loses
        # every movement past the third. So CURRENTY QTY anchors the balance.
        #
        # The withdrawals are still worth keeping — they carry real dates and
        # real TR references — so the lot is opened at
        #     qty_received = CURRENTY QTY + everything we can replay
        # and the withdrawals are replayed against it. The remaining balance
        # then lands on CURRENTY QTY by construction, and the receipt quantity
        # is explicitly a derived figure: what must have arrived for the
        # recorded withdrawals to leave today's balance. Where the sheet's own
        # 'First Qty' disagrees with that, both numbers are reported.
        mode, qty_received, notes = "reconstruct", None, None

        if current_qty is not None and current_qty > 0:
            qty_received = current_qty + out_total
            if first_qty is not None and abs(first_qty - qty_received) > 1e-9:
                issue("warning", "ledger_imbalance", row_no, sku,
                      f"balance anchored on CURRENTY QTY {current_qty:g}; the row's own "
                      f"First Qty {first_qty:g} does not equal that plus its {out_total:g} "
                      f"of withdrawals ({qty_received:g}) — difference "
                      f"{first_qty - qty_received:+g}",
                      {"current_qty": current_qty, "first_qty": first_qty,
                       "withdrawn": out_total, "derived_receipt": qty_received})
                stats["imbalanced"] += 1
            if first_qty is None:
                stats["no_first_qty"] += 1

        elif first_qty is None or first_qty <= 0:
            # No usable receipt. Book what is claimed to be there now.
            if current_qty is not None and current_qty > 0:
                mode, qty_received, outs = "opening", current_qty, []
                if qty_problem:
                    notes = "Opening balance: First Qty unreadable on the source row"
                    stats["opening_damaged"] += 1
                else:
                    notes = "Opening balance: the source row records no receipt event"
                    stats["opening_no_receipt"] += 1
            else:
                issue("error", "unusable_row", row_no, sku,
                      "neither First Qty nor CURRENTY QTY is a usable number — row skipped",
                      {"first_qty": str(row[C_FIRST_QTY])[:60],
                       "current_qty": str(row[C_CURRENT])[:60]})
                stats["skipped"] += 1
                continue

        else:
            # No usable current balance. Fall back to the receipt minus what
            # the sheet says left, which is all that is left to go on.
            if out_total > first_qty + 1e-9:
                # e.g. R2582-BK-TRI-60: 16,159 withdrawn against a 120 receipt.
                issue("error", "withdrawals_exceed_receipt", row_no, sku,
                      f"withdrawals total {out_total:g} against a receipt of {first_qty:g} "
                      f"and there is no CURRENTY QTY to fall back on — withdrawals dropped",
                      {"first_qty": first_qty, "withdrawn": out_total})
                stats["opening_damaged"] += 1
                mode, qty_received, outs = "opening", first_qty, []
                notes = f"Opening balance: source withdrawals ({out_total:g}) were impossible"
            else:
                qty_received = first_qty
                stats["no_current_qty"] += 1

        if qty_received is None or qty_received <= 0:
            stats["skipped"] += 1
            continue

        if not shelf:
            issue("info", "no_shelf", row_no, sku,
                  "no shelf location on this row or any row above it")

        parsed.append({
            "row_ref": row_ref,
            "sku": sku,
            "five_dc": read_5dc(row[C_5DC]),
            "product_name": None,
            "shelf_id": (shelf or "").upper() or None,
            "shelf_text": shelf,
            "pallet_number": clean(row[C_PALLET]),
            "received_on": recv_date.isoformat() if recv_date else None,
            "date_confidence": "exact" if recv_date else "unknown",
            "qty_received": qty_received,
            "source_reference": read_tr(row[C_RECV_TR]),
            "cin7_task_id": None,
            "mode": mode,
            "notes": notes or clean(row[C_NOTES]),
            "outs": outs,
        })
        stats["lots"] += 1
        stats["withdrawals"] += len(outs)
        stats["units_remaining"] += qty_received - sum(o["qty"] for o in outs)
        if recv_date is None:
            stats["undated_lots"] += 1
        if mode == "opening":
            stats["opening_balances"] += 1
        stats["units"] += qty_received

    return parsed, issues, stats


def content_hash(parsed: list[dict]) -> str:
    return hashlib.sha256(
        json.dumps(parsed, sort_keys=True, default=str).encode()).hexdigest()


# ── report ──────────────────────────────────────────────────────────────────
def report(parsed, issues, stats, sb: Supa | None):
    sev = Counter(i["severity"] for i in issues)
    code = Counter(i["code"] for i in issues)
    print()
    print("=" * 72)
    print("  Gateway history import — MAIN Stock Movement")
    print("=" * 72)
    print(f"  rows carrying a SKU      {stats['rows_with_sku']:>8,}")
    print(f"  lots to create           {stats['lots']:>8,}")
    print(f"  withdrawals to replay    {stats['withdrawals']:>8,}")
    print(f"  units on hand after replay {stats['units_remaining']:>6,.0f}   (this is what the screen will show)")
    print(f"  units received (derived)  {stats['units']:>7,.0f}   (on hand + everything replayed out)")
    print(f"  rows skipped entirely    {stats['skipped']:>8,}")
    print()
    print(f"  undated lots             {stats['undated_lots']:>8,}   "
          f"({stats['undated_lots'] / max(stats['lots'], 1):.0%} — FIFO ranks these OLDEST)")
    print(f"  balance from CURRENTY QTY {stats['lots'] - stats['no_current_qty'] - stats['opening_damaged']:>7,}   (the source that matches Cin7 best)")
    print(f"    of which no First Qty  {stats['no_first_qty']:>8,}   (receipt qty derived from balance + withdrawals)")
    print(f"  balance from First Qty   {stats['no_current_qty']:>8,}   (no CURRENTY QTY on the row)")
    print(f"  opening bal: damaged row {stats['opening_damaged']:>8,}   (receipt unreadable or impossible)")
    print(f"  First Qty disagreements  {stats['imbalanced']:>8,}   (both numbers recorded as an issue)")
    print()
    print(f"  issues: {sev['error']} error / {sev['warning']} warning / {sev['info']} info")
    for c, n in code.most_common():
        print(f"      {n:>5,}  {c}")

    print()
    print(f"  SKUs referenced          {len({p['sku'] for p in parsed}):>8,}")
    print(f"  re-cased to match Cin7   {stats['sku_recased']:>8,}   (e.g. 12V-IP20-030W -> 12v-IP20-030w)")
    print(f"  not in cin7_mirror       {stats['sku_unknown']:>8,}   (imported as typed, will read as local-only)")
    print("=" * 72)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    ap.add_argument("--dry-run", action="store_true", help="parse and report, write nothing")
    ap.add_argument("--apply", action="store_true", help="write the ledger")
    ap.add_argument("--replace", action="store_true",
                    help="roll back any previous lot_ledger import first")
    ap.add_argument("--rollback", type=int, metavar="BATCH_ID")
    ap.add_argument("--user", default="excel_migration")
    args = ap.parse_args()

    url, key = _env()
    sb = Supa(url, key)

    if args.rollback:
        print(json.dumps(sb.rpc("gateway_rollback_import",
                                {"p_batch_id": args.rollback, "p_user": args.user}), indent=2))
        return

    if not os.path.exists(args.workbook):
        sys.exit(f"workbook not found: {args.workbook}")

    print("loading the Cin7 product index…")
    sku_index = load_sku_index(sb)
    print(f"  {len(sku_index):,} products")

    print(f"reading {args.workbook}")
    parsed, issues, stats = parse_workbook(args.workbook, sku_index)
    digest = content_hash(parsed)

    if args.dry_run or not args.apply:
        report(parsed, issues, stats, sb)
        print(f"\ncontent hash {digest[:16]}…   (dry run — nothing written)")
        return

    prior = sb.select("gateway_import_batches",
                      f"kind=eq.{KIND}&status=neq.rolled_back&select=id,status,lots_created")
    if prior:
        if not args.replace:
            sys.exit(f"a {KIND} import already exists (batch {prior[0]['id']}, "
                     f"{prior[0]['lots_created']} lots). Re-run with --replace to redo it.")
        for b in prior:
            print(f"rolling back batch {b['id']}…")
            print("   ", sb.rpc("gateway_rollback_import",
                                {"p_batch_id": b["id"], "p_user": args.user}))

    batch = sb.insert("gateway_import_batches", {
        "source_file": os.path.basename(args.workbook), "source_sheet": SHEET,
        "content_hash": digest, "kind": KIND, "status": "running",
        "rows_read": stats["rows_with_sku"], "created_by": args.user,
        "report": {k: v for k, v in stats.items()}})[0]
    batch_id = batch["id"]
    print(f"batch {batch_id} created")

    # One call, one transaction: 771 lots and their movements either all land
    # or none do.
    result = sb.rpc("gateway_import_lot_ledger",
                    {"p_batch_id": batch_id, "p_rows": parsed, "p_user": args.user})
    print("import:", result)

    if issues:
        for i in range(0, len(issues), 400):
            sb.insert("gateway_import_issues",
                      [{**x, "batch_id": batch_id} for x in issues[i:i + 400]])
        print(f"{len(issues)} issues recorded")

    sev = Counter(i["severity"] for i in issues)
    sb.patch("gateway_import_batches", f"id=eq.{batch_id}", {
        "status": "completed", "finished_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "warnings": sev["warning"], "errors": sev["error"],
        "report": {k: v for k, v in stats.items()}})

    report(parsed, issues, stats, sb)
    print(f"\nbatch {batch_id} completed. Review the warnings:")
    print(f"  {url}/rest/v1/gateway_import_issues?batch_id=eq.{batch_id}&resolved=eq.false")


if __name__ == "__main__":
    main()
